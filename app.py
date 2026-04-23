#!/usr/bin/env python3
"""
Flipkart Label Sorter
=====================
Sorts Flipkart shipping label PDFs by product/pack-size,
crops to label-only (no invoice), and produces per-account
sorted-labels PDF + summary PDF.

Run:  python3 app.py
Open: http://localhost:5050
"""

import os, re, io, json, tempfile, traceback, shutil, datetime, gc
from collections import defaultdict
from pathlib import Path

import openpyxl

from flask import Flask, request, jsonify, send_file, Response
from pypdf import PdfReader, PdfWriter
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import mm

IST = datetime.timezone(datetime.timedelta(hours=5, minutes=30))

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 80 * 1024 * 1024  # 80 MB upload cap — protects against OOM on Railway free tier

# Persistent master SKU file — stored in /data (Railway Volume) so it survives restarts
# Falls back to app directory for local development
_data_dir = '/data' if os.path.isdir('/data') else os.path.dirname(__file__)
MASTER_SKU_PATH = os.path.join(_data_dir, 'master_sku.xlsx')

# Persistent output directory — stores latest sorted PDFs per account
OUTPUT_DIR = os.path.join(_data_dir, 'outputs')
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUTS_META = os.path.join(_data_dir, 'outputs_meta.json')  # timestamp + stats per account

KNOWN_ACCOUNTS = ['REFRESHWAVE', 'EONSPARK', 'CUTEST CLUB', 'HELLOHI']
ORDER_DB_PATH  = os.path.join(_data_dir, 'order_db.json')  # persistent order ID history
ORDER_DB_TTL   = 72 * 3600  # 3 days in seconds

# ─────────────────────────────────────────────
# FLIPKART API — TOKEN MANAGER
# ─────────────────────────────────────────────
# Credentials live in Railway env vars — never hardcoded.
# One account supported so far: CUTEST CLUB (Eonspark creds pending).
# Add more accounts by setting FK_<ACCOUNT>_APP_ID / FK_<ACCOUNT>_APP_SECRET.
# Account name key: spaces replaced with underscores, uppercased.
#   e.g. "CUTEST CLUB" → FK_CUTEST_CLUB_APP_ID

FK_TOKEN_CACHE_PATH = os.path.join(_data_dir, 'fk_tokens.json')
FK_API_BASE         = 'https://api.flipkart.net'
FK_REFRESH_BUFFER   = 24 * 3600   # refresh token if less than 24 h remaining

def _fk_env_key(account):
    """Convert account name to env var prefix. 'CUTEST CLUB' → 'FK_CUTEST_CLUB'"""
    return 'FK_' + account.upper().replace(' ', '_')

def _fk_credentials(account):
    """Return (app_id, app_secret) from env vars, or (None, None) if not set."""
    prefix = _fk_env_key(account)
    return os.environ.get(f'{prefix}_APP_ID'), os.environ.get(f'{prefix}_APP_SECRET')

def _fk_load_token_cache():
    if not os.path.exists(FK_TOKEN_CACHE_PATH):
        return {}
    try:
        with open(FK_TOKEN_CACHE_PATH) as f:
            return json.load(f)
    except Exception:
        return {}

def _fk_save_token_cache(cache):
    tmp = FK_TOKEN_CACHE_PATH + '.tmp'
    with open(tmp, 'w') as f:
        json.dump(cache, f)
    os.replace(tmp, FK_TOKEN_CACHE_PATH)

def fk_get_token(account):
    """
    Return a valid access token for the given account.
    Fetches a new one if none cached or within 24 h of expiry.
    Raises RuntimeError if credentials are not configured.
    """
    import requests as _req, base64 as _b64

    app_id, app_secret = _fk_credentials(account)
    if not app_id or not app_secret:
        prefix = _fk_env_key(account)
        raise RuntimeError(
            f'Flipkart credentials not configured for {account}. '
            f'Set {prefix}_APP_ID and {prefix}_APP_SECRET in Railway env vars.'
        )

    cache = _fk_load_token_cache()
    entry = cache.get(account, {})
    now   = datetime.datetime.now(tz=datetime.timezone.utc).timestamp()

    # Return cached token if still valid with buffer
    if entry.get('access_token') and entry.get('expires_at', 0) - now > FK_REFRESH_BUFFER:
        return entry['access_token']

    # Fetch fresh token
    creds = _b64.b64encode(f'{app_id}:{app_secret}'.encode()).decode()
    resp  = _req.get(
        f'{FK_API_BASE}/oauth-service/oauth/token',
        params={'grant_type': 'client_credentials', 'scope': 'Seller_Api'},
        headers={'Authorization': f'Basic {creds}'},
        timeout=15,
    )
    if resp.status_code != 200:
        raise RuntimeError(f'Flipkart token fetch failed [{resp.status_code}]: {resp.text[:200]}')

    data        = resp.json()
    access_token = data['access_token']
    expires_in   = int(data.get('expires_in', 3600))

    # Persist to cache
    cache[account] = {
        'access_token': access_token,
        'expires_at':   now + expires_in,
        'fetched_at':   datetime.datetime.now(tz=IST).strftime('%d %b %Y, %H:%M IST'),
    }
    _fk_save_token_cache(cache)
    print(f'[FK Token] Refreshed token for {account}, expires in {expires_in//3600}h')
    return access_token

def fk_api_get(account, path, params=None):
    """Make an authenticated GET request to the Flipkart seller API."""
    import requests as _req
    token = fk_get_token(account)
    resp  = _req.get(
        f'{FK_API_BASE}{path}',
        params=params or {},
        headers={
            'Authorization': f'Bearer {token}',
            'Content-Type':  'application/json',
        },
        timeout=30,
    )
    return resp

def fk_api_post(account, path, payload=None):
    """Make an authenticated POST request to the Flipkart seller API."""
    import requests as _req
    token = fk_get_token(account)
    resp  = _req.post(
        f'{FK_API_BASE}{path}',
        json=payload or {},
        headers={
            'Authorization': f'Bearer {token}',
            'Content-Type':  'application/json',
        },
        timeout=30,
    )
    return resp

def _atomic_copy(src, dst):
    """Copy src → dst atomically using a temp file + rename to avoid partial writes."""
    dst_tmp = dst + '.tmp'
    shutil.copy2(src, dst_tmp)
    os.replace(dst_tmp, dst)

# ─────────────────────────────────────────────
# ORDER ID DATABASE  (72-hour TTL)
# ─────────────────────────────────────────────

def _load_order_db():
    """Load order DB, purging entries older than 72 hours."""
    if not os.path.exists(ORDER_DB_PATH):
        return {}
    try:
        with open(ORDER_DB_PATH, 'r') as f:
            db = json.load(f)
    except Exception:
        return {}
    now = datetime.datetime.now(tz=datetime.timezone.utc).timestamp()
    purged = {}
    for account, orders in db.items():
        kept = {}
        for oid, val in orders.items():
            if isinstance(val, dict):
                # New format: {awb: timestamp}
                kept_awbs = {awb: ts for awb, ts in val.items()
                             if now - ts <= ORDER_DB_TTL}
                if kept_awbs:
                    kept[oid] = kept_awbs
            else:
                # Legacy format: flat timestamp
                if now - val <= ORDER_DB_TTL:
                    kept[oid] = val
        if kept:
            purged[account] = kept
    return purged

def _save_order_db(db):
    tmp = ORDER_DB_PATH + '.tmp'
    with open(tmp, 'w') as f:
        json.dump(db, f)
    os.replace(tmp, ORDER_DB_PATH)

def _record_order_ids(account_order_ids):
    """Add newly sorted order IDs to the DB.
    account_order_ids: {account: [(oid, awb), ...]}
    Stored as {account: {oid: {awb: timestamp}}}
    """
    db  = _load_order_db()
    now = datetime.datetime.now(tz=datetime.timezone.utc).timestamp()
    for account, keys in account_order_ids.items():
        if account not in db:
            db[account] = {}
        for item in keys:
            if isinstance(item, (list, tuple)) and len(item) == 2:
                oid, awb = item
            else:
                oid, awb = item, ''
            if oid not in db[account]:
                db[account][oid] = {}
            # Handle legacy flat format
            if not isinstance(db[account][oid], dict):
                db[account][oid] = {}
            db[account][oid][awb] = now
    _save_order_db(db)

def extract_order_ids(text):
    """Extract all OD... order IDs from a label page."""
    return re.findall(r'OD\d{15,}', text)

def extract_awb(text):
    """Extract AWB/tracking ID (FMP...) from a label page."""
    m = re.search(r'(?:DTr:|AWB[^\w]*)?(FMP[A-Z]\d+)', text)
    return m.group(1) if m else ''

def extract_label_keys(text):
    """Return list of (order_id, awb) composite keys for a label page.
    Two labels with same OID but different AWB are treated as distinct."""
    oids = extract_order_ids(text)
    awb  = extract_awb(text)
    return [(oid, awb) for oid in oids]

# ─────────────────────────────────────────────
# TELEGRAM BOT
# ─────────────────────────────────────────────
import urllib.request as _urllib_req

TELEGRAM_TOKEN     = '8734907502:AAF2qgG1eILANUS-VxZrUrM6GfudQi71qCc'
TELEGRAM_OWNER     = 530170157   # Only this chat_id can use commands
TELEGRAM_API       = f'https://api.telegram.org/bot{TELEGRAM_TOKEN}'
# To broadcast to a team group later, add the group chat_id here:
TELEGRAM_BROADCAST = [TELEGRAM_OWNER]

def _tg_send_message(chat_id, text):
    """Send a plain HTML text message. Silent on failure."""
    try:
        payload = json.dumps({
            'chat_id': chat_id, 'text': text, 'parse_mode': 'HTML'
        }).encode()
        _urllib_req.urlopen(
            _urllib_req.Request(
                f'{TELEGRAM_API}/sendMessage', data=payload,
                headers={'Content-Type': 'application/json'}
            ), timeout=30
        )
    except Exception as e:
        print(f'[Telegram] sendMessage failed: {e}')

def _tg_send_document(chat_id, file_path, caption=''):
    """Send a PDF file. Uses requests library for multipart upload."""
    try:
        import requests as _req
        with open(file_path, 'rb') as f:
            _req.post(
                f'{TELEGRAM_API}/sendDocument',
                data={'chat_id': chat_id, 'caption': caption, 'parse_mode': 'HTML'},
                files={'document': (os.path.basename(file_path), f, 'application/pdf')},
                timeout=60
            )
    except Exception as e:
        print(f'[Telegram] sendDocument failed: {e}')

def tg_notify_sort_done(account, total, normal_count, mixed_count, unknown_count,
                        labels_path, summary_path):
    """Broadcast sort-complete alert + both PDFs to all targets."""
    timestamp = datetime.datetime.now(tz=IST).strftime('%d %b %Y, %H:%M IST')
    msg = (
        f'\U0001f3f7\ufe0f <b>Sort Complete — {account}</b>\n'
        f'\U0001f550 {timestamp}\n\n'
        f'\U0001f4e6 Total labels: <b>{total}</b>\n'
        f'\u2705 Identified: <b>{normal_count}</b>\n'
        f'\u26a1 Mixed orders: <b>{mixed_count}</b>\n'
        f'\u2753 Unidentified: <b>{unknown_count}</b>'
    )
    for chat_id in TELEGRAM_BROADCAST:
        _tg_send_message(chat_id, msg)
        _tg_send_document(chat_id, labels_path,
                          caption=f'\U0001f4c4 Sorted Labels — {account}')
        _tg_send_document(chat_id, summary_path,
                          caption=f'\U0001f4ca Summary — {account}')

def tg_handle_command(chat_id, text):
    """Handle bot commands. Only TELEGRAM_OWNER can interact."""
    if chat_id != TELEGRAM_OWNER:
        return  # silently ignore everyone else
    text = (text or '').strip()

    if text.startswith('/start'):
        _tg_send_message(chat_id,
            '\U0001f44b <b>Flipkart Ops Bot</b>\n\n'
            "I'll notify you after every sort with both PDFs automatically.\n\n"
            '<b>Commands:</b>\n'
            '/status — latest sort info for all accounts\n'
            '/download refreshwave\n'
            '/download eonspark\n'
            '/download cutest_club\n'
            '/download hellohi'
        )

    elif text.startswith('/status'):
        meta = {}
        if os.path.exists(OUTPUTS_META):
            with open(OUTPUTS_META) as f:
                try: meta = json.load(f)
                except: pass
        if not meta:
            _tg_send_message(chat_id, '\U0001f4ed No sorts run yet.')
            return
        lines = ['\U0001f4ca <b>Latest Sort Status</b>\n']
        for acc in KNOWN_ACCOUNTS:
            if acc in meta:
                m = meta[acc]
                lines.append(
                    f'<b>{acc}</b>\n'
                    f'  \U0001f550 {m["timestamp"]}\n'
                    f'  \U0001f4e6 {m["total"]} labels | \u2705 {m["sku_count"]} identified\n'
                )
            else:
                lines.append(f'<b>{acc}</b> — no data yet\n')
        _tg_send_message(chat_id, '\n'.join(lines))

    elif text.startswith('/download'):
        parts = text.split(maxsplit=1)
        if len(parts) < 2:
            _tg_send_message(chat_id,
                '\u26a0\ufe0f Usage: /download &lt;account&gt;\nE.g. /download refreshwave')
            return
        query = parts[1].strip().upper().replace('_', ' ')
        matched = next((a for a in KNOWN_ACCOUNTS if query in a or a in query), None)
        if not matched:
            _tg_send_message(chat_id,
                f'\u274c Unknown account: {query}\n'
                f'Known: {", ".join(KNOWN_ACCOUNTS)}')
            return
        safe_name    = re.sub(r'[^A-Za-z0-9_]', '_', matched)
        labels_path  = os.path.join(OUTPUT_DIR, f'{safe_name}_labels.pdf')
        summary_path = os.path.join(OUTPUT_DIR, f'{safe_name}_summary.pdf')
        if not os.path.exists(labels_path):
            _tg_send_message(chat_id, f'\U0001f4ed No files found for {matched} yet.')
            return
        _tg_send_message(chat_id, f'\U0001f4e4 Sending latest files for <b>{matched}</b>\u2026')
        _tg_send_document(chat_id, labels_path,  caption=f'\U0001f4c4 Sorted Labels — {matched}')
        _tg_send_document(chat_id, summary_path, caption=f'\U0001f4ca Summary — {matched}')

    else:
        _tg_send_message(chat_id,
            '\U0001f916 Unknown command.\nTry /start, /status, or /download &lt;account&gt;')


# ─────────────────────────────────────────────
# HTML FRONTEND
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

ACCOUNT_PATTERNS = {
    "EONSPARK": "EONSPARK",
    "REFRESHWAVE": "REFRESHWAVE",
    "CUTEST CLUB": "CUTEST CLUB",
    "HELLOHI": "HELLOHI",
}

def detect_account(text):
    """Detect seller account from label text."""
    text_upper = text.upper()
    for key, name in ACCOUNT_PATTERNS.items():
        if key in text_upper:
            return name
    # Fallback: extract "Sold By:XXXX" pattern
    m = re.search(r'Sold By[:\s]+([A-Z][A-Za-z0-9 &\(\)]+?)(?:,|\n|PRIVATE|LIMITED)', text, re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()[:40]
    return "UNKNOWN_ACCOUNT"

def extract_skus_from_page(text):
    """
    Extract list of {sku, qty} from a label page.

    pypdf splits the SKU table across multiple lines. The pattern after
    'SKU ID | Description / QTY' is one of two formats:

    Format A — SKU and description on separate lines, qty alone:
        "1"
        "4 QTY - "          ← optional prefix line
        "SKU NAME "
        "| Description text"
        "4"                  ← qty (standalone integer line)
        "FMPC..."            ← AWB → end of table

    Format B — SKU | description on ONE line (mixed orders, 2nd item):
        "2"
        "SKU NAME | Description"
        "1"
        "FMPC..."

    Strategy:
      1. Find the header line.
      2. Collect all lines until the AWB/FMP line.
      3. Reconstruct SKU entries by detecting row-number lines as boundaries.
    """
    lines = [l.strip() for l in text.split('\n')]
    skus = []

    # Find start of SKU table
    start = None
    for i, line in enumerate(lines):
        if 'SKU ID' in line and 'Description' in line:
            start = i + 1
            break
    if start is None:
        return skus

    # Skip the 'QTY' header line if present
    if start < len(lines) and lines[start].strip() == 'QTY':
        start += 1

    # Collect table lines until AWB/FMP barcode line (end of label section)
    table_lines = []
    for line in lines[start:]:
        if re.match(r'^(FMP[A-Z0-9]|FMPP|AWB|Tax Invoice)', line):
            break
        table_lines.append(line)

    # Split into per-row chunks using standalone row-number lines as boundaries
    # A row number line is a bare integer (1, 2, 3...) NOT a qty-looking line
    # We detect it as: digit(s) only, and it appears before any '|' in the chunk
    row_chunks = []
    current = []
    for line in table_lines:
        if re.match(r'^\d+$', line) and not current:
            # This is the row number starting a new SKU entry
            current = [line]
        elif re.match(r'^\d+$', line) and current:
            # Could be qty (end of current entry) or next row number
            # If current already has a '|' containing line, this is qty → close chunk
            chunk_text = ' '.join(current)
            if '|' in chunk_text:
                current.append(line)  # qty line
                row_chunks.append(current)
                current = []
            else:
                # No | yet — ambiguous, treat as qty and close
                current.append(line)
                row_chunks.append(current)
                current = []
        else:
            if current:
                current.append(line)
    if current:
        row_chunks.append(current)

    for chunk in row_chunks:
        if not chunk:
            continue
        # Join all lines, find SKU (part before first |) and qty (last standalone int)
        full = ' '.join(chunk)

        # Remove leading row number
        full = re.sub(r'^\d+\s*', '', full, count=1)

        # Remove optional "N QTY - " prefix (e.g. "4 QTY - ")
        full = re.sub(r'^\d+\s*QTY\s*-\s*', '', full, flags=re.IGNORECASE)

        # Extract SKU name: everything before the first ' | '
        pipe_match = re.search(r'\s*\|\s*', full)
        if pipe_match:
            sku_name = full[:pipe_match.start()].strip()
        else:
            sku_name = full.strip()

        # Extract qty: last standalone integer in the chunk
        all_ints = re.findall(r'(?<!\S)(\d+)(?!\S)', ' '.join(chunk))
        qty = int(all_ints[-1]) if all_ints else 1

        if sku_name and sku_name.lower() not in ('', 'qty'):
            skus.append({'sku': sku_name, 'qty': qty})

    return skus

def load_sku_master(xlsx_bytes):
    """
    Parse xlsx (multi-sheet) into:
      { SHEET_NAME_UPPER: { sku: {product, pack_size, pack_raw} } }
    Each sheet = one seller account. Sheet name matched against detected
    account name from labels via get_account_master().
    """
    import openpyxl
    master = {}
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    for sheet in wb.worksheets:
        account_key = sheet.title.strip().upper()
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = None
        data_start = 0
        for i, row in enumerate(rows):
            norm = [str(c).strip().lower() if c else '' for c in row]
            if any(x in norm for x in ['sku', 'sku id', 'skuid']):
                header = norm
                data_start = i + 1
                break
        if not header:
            continue
        def col(names, start_after=None):
            # Exact match first
            for n in names:
                for i, h in enumerate(header):
                    if start_after is not None and i <= start_after:
                        continue
                    if h == n:
                        return i
            # Prefix match (e.g. 'productname 1' matches 'productname')
            for n in names:
                for i, h in enumerate(header):
                    if start_after is not None and i <= start_after:
                        continue
                    if h.startswith(n):
                        return i
            return None
        sku_col   = col(['sku', 'sku id', 'skuid'])
        prod1_col = col(['productname 1', 'productname', 'product name', 'product'])
        pack1_col = col(['packsize', 'pack size', 'pack'])
        # ProductName 2 and its PackSize must come AFTER prod1_col in the header
        prod2_col = col(['productname 2', 'productname'], start_after=prod1_col) if prod1_col is not None else None
        pack2_col = col(['packsize', 'pack size', 'pack'], start_after=pack1_col) if pack1_col is not None else None
        if sku_col is None:
            continue
        account_skus = {}
        for row in rows[data_start:]:
            sku      = str(row[sku_col]).strip()    if row[sku_col]                              else ''
            product  = str(row[prod1_col]).strip()  if prod1_col is not None and row[prod1_col] else ''
            pack_raw = str(row[pack1_col]).strip()  if pack1_col is not None and row[pack1_col] else '0'
            prod2    = str(row[prod2_col]).strip()  if prod2_col is not None and row[prod2_col] else ''
            pack2_raw= str(row[pack2_col]).strip()  if pack2_col is not None and row[pack2_col] else '0'
            if sku and sku.lower() not in ('none', ''):
                try:
                    pack_num = int(float(pack_raw)) if pack_raw else 0
                except:
                    pack_num = 0
                try:
                    pack2_num = int(float(pack2_raw)) if pack2_raw else 0
                except:
                    pack2_num = 0
                entry = {'product': product, 'pack_size': pack_num, 'pack_raw': pack_raw,
                         'product2': prod2, 'pack_size2': pack2_num}
                account_skus[sku] = entry
        master[account_key] = account_skus
    return master


def get_account_master(master, account_name):
    """
    Fuzzy-match detected account name against xlsx sheet names.
    e.g. detected "REFRESHWAVE" matches sheet "Refreshwave".
    Returns the SKU dict for that account, or empty dict if no match.
    """
    account_upper = account_name.upper()
    if account_upper in master:
        return master[account_upper]
    for sheet_key in master:
        if sheet_key in account_upper or account_upper in sheet_key:
            return master[sheet_key]
    return {}

def crop_page_top_half(page):
    """Crop PDF page to top half (label only, removes invoice)."""
    mb = page.mediabox
    mid_y = (mb.top + mb.bottom) / 2
    page.mediabox.bottom = mid_y
    page.cropbox.bottom = mid_y
    return page

def is_mixed_order(skus_on_page, master):
    """
    Returns True if this label is a mixed order:
    - Multiple SKUs on one label regardless of whether same/different product
    - Or 1 SKU but qty > 1 with different tracking (handled at merge level)
    """
    if len(skus_on_page) > 1:
        return True
    return False

def classify_pages(pages_data, master):
    """
    pages_data: list of {page_idx, skus: [{sku, qty}], text}
    Returns: (normal_pages, dual_pages, mixed_pages, unknown_pages)
    - normal : single SKU, single product
    - dual   : single SKU, but master says it contains 2 products
    - mixed  : multiple SKU entries on one label
    - unknown: SKU not in master
    """
    normal = []
    dual = []
    mixed = []
    unknown = []

    for pd in pages_data:
        skus = pd['skus']

        if not skus:
            unknown.append(pd)
            continue

        # Multiple SKU lines on label
        if len(skus) > 1:
            # Check if all SKU lines resolve to the same product
            resolved_products = set()
            total_qty = 0
            all_known = True
            for s in skus:
                info = master.get(s['sku'])
                if info and info['product']:
                    resolved_products.add((info['product'], info['pack_size']))
                    total_qty += s['qty']
                else:
                    all_known = False
                    break
            if all_known and len(resolved_products) == 1:
                # All lines are the same product — treat as normal order
                # Effective pack size = base pack_size × total number of units
                product, pack_size = list(resolved_products)[0]
                pd['primary_product'] = product
                pd['pack_size'] = pack_size * total_qty  # e.g. Pack 1 × 2 units = Pack 2
                pd['effective_qty'] = total_qty
                normal.append(pd)
            else:
                mixed.append(pd)
            continue

        # Single SKU — look it up
        sku_name = skus[0]['sku']
        info = master.get(sku_name)
        if not info:
            pd['primary_product'] = sku_name
            pd['pack_size'] = 0
            unknown.append(pd)
            continue

        pd['primary_product'] = info['product']
        pd['pack_size']        = info['pack_size']

        # Check if this SKU is a dual-product bundle
        if info.get('product2'):
            pd['product2']   = info['product2']
            pd['pack_size2'] = info['pack_size2']
            dual.append(pd)
        else:
            normal.append(pd)

    return normal, dual, mixed, unknown

def sort_normal(pages):
    """
    Sort normal pages to match summary order:
    - Products ordered by total label count desc
    - Within each product, pack_size desc
    """
    from collections import defaultdict
    # Count total labels per product
    prod_counts = defaultdict(int)
    for p in pages:
        prod_counts[p.get('primary_product', '')] += 1
    return sorted(pages, key=lambda p: (
        -prod_counts[p.get('primary_product', '')],   # product count desc
        p.get('primary_product', '').lower(),          # then alpha (tiebreak)
        -p.get('pack_size', 0)                         # then pack size desc
    ))

def build_sorted_pdf(all_page_indices, reader, output_path, crop=True):
    """Build PDF from page indices, optionally cropping to top half."""
    writer = PdfWriter()
    for idx in all_page_indices:
        page = reader.pages[idx]
        if crop:
            page = crop_page_top_half(page)
        writer.add_page(page)
    with open(output_path, 'wb') as f:
        writer.write(f)

def build_summary_pdf(account_name, normal_pages, dual_pages, mixed_pages, unknown_pages, master, output_path):
    """Build summary PDF table."""
    # Count by product+packsize for normal pages
    product_counts = defaultdict(int)
    for p in normal_pages:
        key = (p.get('primary_product',''), p.get('pack_size', 0))
        product_counts[key] += 1

    # Mixed counts
    mixed_count = len(mixed_pages)
    unknown_count = len(unknown_pages)

    doc = SimpleDocTemplate(output_path, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"Label Sort Summary — {account_name}", styles['Title']))
    story.append(Spacer(1, 4*mm))

    dual_count = len(dual_pages)
    total = len(normal_pages) + dual_count + mixed_count + unknown_count
    story.append(Paragraph(
        f"Total labels: {total}  |  Single-product: {len(normal_pages)}  |  Dual-product bundles: {dual_count}  |  Mixed: {mixed_count}  |  Unidentified: {unknown_count}",
        styles['Normal']
    ))
    story.append(Spacer(1, 6*mm))

    # Table data
    data = [["#", "Product", "Pack Size", "Labels"]]

    # Sorted product rows
    # Re-sort by count desc within same product
    from collections import OrderedDict
    # Group by product, sort pack desc, then by count
    prod_group = defaultdict(list)
    for (prod, pack), cnt in product_counts.items():
        prod_group[prod].append((pack, cnt))

    # Sort products by total count desc
    prod_totals = {p: sum(c for _, c in items) for p, items in prod_group.items()}
    row_num = 1
    for prod in sorted(prod_group, key=lambda p: -prod_totals[p]):
        for pack, cnt in sorted(prod_group[prod], key=lambda x: -x[0]):
            pack_label = f"Pack {pack}" if pack else "—"
            data.append([row_num, prod, pack_label, cnt])
            row_num += 1

    # Dual-product bundle rows
    dual_counts = defaultdict(int)
    dual_labels = {}  # key -> display string
    for p in dual_pages:
        p1 = p.get('primary_product', '')
        ps1 = p.get('pack_size', 0)
        p2 = p.get('product2', '')
        ps2 = p.get('pack_size2', 0)
        key = (p1, ps1, p2, ps2)
        dual_counts[key] += 1
        dual_labels[key] = f"{p1} (Pack {ps1}) + {p2} (Pack {ps2})"
    for key, cnt in sorted(dual_counts.items(), key=lambda x: -x[1]):
        data.append([row_num, f"📦 {dual_labels[key]}", "Bundle", cnt])
        row_num += 1

    # Unknown rows
    unknown_skus = defaultdict(int)
    for p in unknown_pages:
        sku = p['skus'][0]['sku'] if p['skus'] else 'Unknown'
        unknown_skus[sku] += 1
    for sku, cnt in sorted(unknown_skus.items(), key=lambda x: -x[1]):
        data.append([row_num, f"❓ {sku}", "—", cnt])
        row_num += 1

    table = Table(data, colWidths=[12*mm, 110*mm, 28*mm, 20*mm])
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2874f0')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('ALIGN', (3,0), (3,-1), 'CENTER'),
        ('FONTSIZE', (0,1), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#cccccc')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
    ])

    # Highlight top products (count >= 5)
    for i, row in enumerate(data[1:], 1):
        if isinstance(row[3], int) and row[3] >= 5 and not str(row[1]).startswith('❓'):
            style.add('BACKGROUND', (0,i), (-1,i), colors.HexColor('#d5e8d4'))
            style.add('FONTNAME', (0,i), (-1,i), 'Helvetica-Bold')
        if str(row[1]).startswith('📦'):
            style.add('BACKGROUND', (0,i), (-1,i), colors.HexColor('#e8f4fd'))
            style.add('FONTNAME', (0,i), (-1,i), 'Helvetica-Bold')
        if str(row[1]).startswith('❓'):
            style.add('BACKGROUND', (0,i), (-1,i), colors.HexColor('#f8d7da'))

    table.setStyle(style)
    story.append(table)

    # ── Mixed Orders Detail Section ──
    if mixed_pages:
        story.append(Spacer(1, 8*mm))
        story.append(Paragraph("⚡ Mixed Orders Detail", styles['Heading2']))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(
            "Each row below is one label. Pack the items listed together in one shipment.",
            styles['Normal']
        ))
        story.append(Spacer(1, 4*mm))

        mixed_data = [["#", "Product", "Qty"]]
        for idx, mp in enumerate(mixed_pages, 1):
            skus = mp.get('skus', [])
            # Consolidate duplicate SKUs: sum quantities, show product name
            consolidated = {}  # sku -> {label, qty}
            for s in skus:
                sku_key = s['sku']
                info = master.get(sku_key)
                if info and info['product']:
                    display = f"{info['product']} (Pack {info['pack_size']})"
                else:
                    display = sku_key
                if sku_key in consolidated:
                    consolidated[sku_key]['qty'] += s['qty']
                else:
                    consolidated[sku_key] = {'label': display, 'qty': s['qty']}
            sku_lines = [v['label'] for v in consolidated.values()]
            qty_lines = [str(v['qty']) for v in consolidated.values()]
            mixed_data.append([idx, "\n".join(sku_lines), "\n".join(qty_lines)])

        mixed_table = Table(mixed_data, colWidths=[12*mm, 138*mm, 20*mm])
        mixed_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e67e22')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ALIGN', (2,0), (2,-1), 'CENTER'),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#fffbf0'), colors.HexColor('#fff3cd')]),
            ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#cccccc')),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ])
        mixed_table.setStyle(mixed_style)
        story.append(mixed_table)

    doc.build(story)


ACCOUNT_ORDER = ['REFRESHWAVE', 'EONSPARK', 'HELLOHI', 'CUTEST CLUB']

def build_consolidated_pdf(all_account_data, output_path):
    """
    Build a single consolidated sorted-labels PDF across all accounts.

    all_account_data: {account: {'normal': [...], 'dual': [...], 'mixed': [...], 'unknown': [...]}}

    Sort order:
      1. Product group total (sum of ALL pack sizes across ALL accounts) DESC
      2. Within product: pack size DESC
      3. Within product+pack: fixed account order (ACCOUNT_ORDER)
    """
    # ── Gather all (product, pack_size) combinations and compute totals ────────
    # product_total: {product_name: total_label_count_across_all_packs_and_accounts}
    # pack_data: {(product, pack_size): {account: [page_dicts]}}
    product_total = defaultdict(int)
    pack_data     = defaultdict(lambda: defaultdict(list))  # (prod,pack) -> {acct: [pages]}

    for account, buckets in all_account_data.items():
        for page in buckets['normal']:
            prod = page.get('primary_product', '')
            pack = page.get('pack_size', 0)
            product_total[prod] += 1
            pack_data[(prod, pack)][account].append(page)
        for page in buckets['dual']:
            prod = page.get('primary_product', '')
            pack = page.get('pack_size', 0)
            product_total[prod] += 1
            pack_data[(prod, pack)][account].append(page)

    # ── Build ordered list of pages ────────────────────────────────────────────
    ordered_pages = []

    # Sort products by total DESC, then alpha for tiebreak
    sorted_products = sorted(
        set(p for p, _ in pack_data.keys()),
        key=lambda p: (-product_total[p], p.lower())
    )

    for prod in sorted_products:
        # Get all pack sizes for this product, sort DESC
        packs = sorted(
            set(pack for (p, pack) in pack_data.keys() if p == prod),
            reverse=True
        )
        for pack in packs:
            acct_map = pack_data[(prod, pack)]
            # Within each product+pack, emit pages in fixed account order
            for account in ACCOUNT_ORDER:
                for page in acct_map.get(account, []):
                    ordered_pages.append(page)
            # Any account not in ACCOUNT_ORDER (unknown accounts) appended last
            for account, pages in acct_map.items():
                if account not in ACCOUNT_ORDER:
                    ordered_pages.extend(pages)

    # Append unknown and mixed from all accounts at the end
    for account in ACCOUNT_ORDER:
        buckets = all_account_data.get(account, {})
        ordered_pages.extend(buckets.get('unknown', []))
    for account in ACCOUNT_ORDER:
        buckets = all_account_data.get(account, {})
        ordered_pages.extend(buckets.get('mixed', []))

    # ── Write PDF ──────────────────────────────────────────────────────────────
    pages_by_pdf = defaultdict(list)
    for pos, pd_item in enumerate(ordered_pages):
        pages_by_pdf[pd_item['orig_path']].append((pos, pd_item['orig_idx']))

    page_slots = [None] * len(ordered_pages)
    open_readers = {}
    for pdf_src, idx_pairs in pages_by_pdf.items():
        reader = PdfReader(pdf_src)
        open_readers[pdf_src] = reader
        for pos, page_idx in idx_pairs:
            page_slots[pos] = reader.pages[page_idx]

    writer = PdfWriter()
    for page in page_slots:
        if page is not None:
            writer.add_page(page)
    with open(output_path, 'wb') as f:
        writer.write(f)

    del writer, page_slots, open_readers
    gc.collect()


def build_consolidated_summary_pdf(all_account_data, output_path):
    """
    Build consolidated summary PDF with a per-account breakdown table.

    Columns: # | Product | Pack Size | Refreshwave | Eonspark | Hellohi | Cutest Club | Total
    Rows sorted by product-level total DESC, then pack size DESC within product.
    """
    # Gather counts per (product, pack) per account
    pack_counts = defaultdict(lambda: defaultdict(int))  # (prod,pack) -> {account: count}
    product_total = defaultdict(int)

    for account, buckets in all_account_data.items():
        for page in buckets['normal'] + buckets['dual']:
            prod = page.get('primary_product', '')
            pack = page.get('pack_size', 0)
            pack_counts[(prod, pack)][account] += 1
            product_total[prod] += 1

    # Mixed / unknown totals
    mixed_total   = sum(len(b.get('mixed',   [])) for b in all_account_data.values())
    unknown_total = sum(len(b.get('unknown', [])) for b in all_account_data.values())
    grand_total   = sum(product_total.values()) + mixed_total + unknown_total

    doc = SimpleDocTemplate(output_path, pagesize=A4,
        leftMargin=10*mm, rightMargin=10*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story  = []

    story.append(Paragraph('Consolidated Label Sort Summary — All Accounts', styles['Title']))
    story.append(Spacer(1, 3*mm))
    timestamp = datetime.datetime.now(tz=IST).strftime('%d %b %Y, %H:%M IST')
    story.append(Paragraph(
        f'Generated: {timestamp} &nbsp;|&nbsp; '
        f'Total labels: <b>{grand_total}</b> &nbsp;|&nbsp; '
        f'Mixed: <b>{mixed_total}</b> &nbsp;|&nbsp; '
        f'Unidentified: <b>{unknown_total}</b>',
        styles['Normal']
    ))
    story.append(Spacer(1, 6*mm))

    # Table header
    acct_labels = ['RW', 'ES', 'HH', 'CC']  # short column headers
    header_row  = ['#', 'Product', 'Pack', ] + acct_labels + ['Total']
    data = [header_row]

    # Sort products by total DESC
    sorted_products = sorted(
        set(p for p, _ in pack_counts.keys()),
        key=lambda p: (-product_total[p], p.lower())
    )

    row_num = 1
    for prod in sorted_products:
        packs = sorted(
            set(pack for (p, pack) in pack_counts.keys() if p == prod),
            reverse=True
        )
        for pack in packs:
            acct_map  = pack_counts[(prod, pack)]
            row_total = sum(acct_map.values())
            pack_label = f'Pack {pack}' if pack else '—'
            row = [row_num, prod, pack_label]
            for account in ACCOUNT_ORDER:
                cnt = acct_map.get(account, 0)
                row.append(cnt if cnt else '—')
            row.append(row_total)
            data.append(row)
            row_num += 1

    # Mixed / unknown — summary count rows in main table
    if mixed_total:
        data.append([row_num, '⚡ Mixed Orders', '—', '—', '—', '—', '—', mixed_total])
        row_num += 1
    if unknown_total:
        data.append([row_num, '❓ Unidentified', '—', '—', '—', '—', '—', unknown_total])

    col_widths = [10*mm, 75*mm, 18*mm, 18*mm, 18*mm, 18*mm, 18*mm, 18*mm]
    table = Table(data, colWidths=col_widths)

    style = TableStyle([
        ('BACKGROUND',   (0,0), (-1,0),  colors.HexColor('#1a237e')),
        ('TEXTCOLOR',    (0,0), (-1,0),  colors.white),
        ('FONTNAME',     (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,0),  9),
        ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
        ('ALIGN',        (1,0), (1,-1),  'LEFT'),
        ('FONTSIZE',     (0,1), (-1,-1), 8),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, colors.HexColor('#f0f4ff')]),
        ('GRID',         (0,0), (-1,-1), 0.4, colors.HexColor('#bbbbbb')),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 4),
        ('LEFTPADDING',  (0,0), (-1,-1), 4),
        ('FONTNAME',     (-1,1),(-1,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR',    (-1,1),(-1,-1), colors.HexColor('#1a237e')),
    ])
    for i, row in enumerate(data[1:], 1):
        total_val = row[-1]
        if isinstance(total_val, int) and total_val >= 10:
            style.add('BACKGROUND', (0,i), (-1,i), colors.HexColor('#e8eaf6'))
            style.add('FONTNAME',   (0,i), (-1,i), 'Helvetica-Bold')
    table.setStyle(style)
    story.append(table)

    # Legend
    story.append(Spacer(1, 5*mm))
    story.append(Paragraph(
        '<b>Columns:</b> RW = Refreshwave &nbsp;|&nbsp; ES = Eonspark &nbsp;|&nbsp; '
        'HH = Hellohi &nbsp;|&nbsp; CC = Cutest Club',
        styles['Normal']
    ))

    # ── Mixed Orders Detail (with Account column) ─────────────────────────────
    # Collect all mixed pages across accounts in fixed account order
    all_mixed = []
    for account in ACCOUNT_ORDER:
        for page in all_account_data.get(account, {}).get('mixed', []):
            all_mixed.append((account, page))
    # Also any unknown accounts
    for account, buckets in all_account_data.items():
        if account not in ACCOUNT_ORDER:
            for page in buckets.get('mixed', []):
                all_mixed.append((account, page))

    if all_mixed:
        story.append(Spacer(1, 8*mm))
        story.append(Paragraph('⚡ Mixed Orders Detail', styles['Heading2']))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(
            'Each row below is one label. Pack the items listed together in one shipment.',
            styles['Normal']
        ))
        story.append(Spacer(1, 4*mm))

        # Build per-account master lookup for display names
        full_master = {}
        if os.path.exists(MASTER_SKU_PATH):
            with open(MASTER_SKU_PATH, 'rb') as _f:
                full_master = load_sku_master(_f.read())

        mixed_data = [['#', 'Account', 'Product', 'Qty']]
        for idx, (account, mp) in enumerate(all_mixed, 1):
            acct_master = get_account_master(full_master, account)
            skus = mp.get('skus', [])
            consolidated_skus = {}
            for s in skus:
                info = acct_master.get(s['sku'])
                display = f"{info['product']} (Pack {info['pack_size']})" if info and info['product'] else s['sku']
                if s['sku'] in consolidated_skus:
                    consolidated_skus[s['sku']]['qty'] += s['qty']
                else:
                    consolidated_skus[s['sku']] = {'label': display, 'qty': s['qty']}
            sku_lines = [v['label'] for v in consolidated_skus.values()]
            qty_lines  = [str(v['qty'])  for v in consolidated_skus.values()]
            mixed_data.append([idx, account, '\n'.join(sku_lines), '\n'.join(qty_lines)])

        mixed_table = Table(mixed_data, colWidths=[10*mm, 32*mm, 106*mm, 17*mm])
        mixed_style = TableStyle([
            ('BACKGROUND',    (0,0), (-1,0),  colors.HexColor('#e67e22')),
            ('TEXTCOLOR',     (0,0), (-1,0),  colors.white),
            ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,0),  9),
            ('ALIGN',         (0,0), (-1,-1), 'LEFT'),
            ('ALIGN',         (3,0), (3,-1),  'CENTER'),
            ('FONTSIZE',      (0,1), (-1,-1), 8),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.HexColor('#fffbf0'), colors.HexColor('#fff3cd')]),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor('#cccccc')),
            ('TOPPADDING',    (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING',   (0,0), (-1,-1), 6),
            ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ])
        mixed_table.setStyle(mixed_style)
        story.append(mixed_table)

    # ── Unidentified Detail (with Account column) ─────────────────────────────
    all_unknown = []
    for account in ACCOUNT_ORDER:
        for page in all_account_data.get(account, {}).get('unknown', []):
            all_unknown.append((account, page))
    for account, buckets in all_account_data.items():
        if account not in ACCOUNT_ORDER:
            for page in buckets.get('unknown', []):
                all_unknown.append((account, page))

    if all_unknown:
        story.append(Spacer(1, 8*mm))
        story.append(Paragraph('❓ Unidentified SKUs Detail', styles['Heading2']))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(
            'These labels could not be matched to the Master SKU file. Add them to the SKU editor to resolve.',
            styles['Normal']
        ))
        story.append(Spacer(1, 4*mm))

        # Group by (account, sku) for a cleaner count table
        unknown_counts = defaultdict(int)   # (account, sku) -> count
        for account, page in all_unknown:
            sku = page['skus'][0]['sku'] if page.get('skus') else 'Unknown'
            unknown_counts[(account, sku)] += 1

        unk_data = [['#', 'Account', 'SKU', 'Labels']]
        for i, ((account, sku), cnt) in enumerate(
            sorted(unknown_counts.items(), key=lambda x: (
                ACCOUNT_ORDER.index(x[0][0]) if x[0][0] in ACCOUNT_ORDER else 99,
                -x[1]
            )), 1
        ):
            unk_data.append([i, account, sku, cnt])

        unk_table = Table(unk_data, colWidths=[10*mm, 32*mm, 117*mm, 16*mm])
        unk_style = TableStyle([
            ('BACKGROUND',    (0,0), (-1,0),  colors.HexColor('#c0392b')),
            ('TEXTCOLOR',     (0,0), (-1,0),  colors.white),
            ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,0),  9),
            ('ALIGN',         (0,0), (-1,-1), 'LEFT'),
            ('ALIGN',         (3,0), (3,-1),  'CENTER'),
            ('FONTSIZE',      (0,1), (-1,-1), 8),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.HexColor('#fdf2f2'), colors.HexColor('#f8d7da')]),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor('#cccccc')),
            ('TOPPADDING',    (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('LEFTPADDING',   (0,0), (-1,-1), 6),
        ])
        unk_table.setStyle(unk_style)
        story.append(unk_table)

    doc.build(story)


# ─────────────────────────────────────────────
# SALES ANALYTICS  helpers
# ─────────────────────────────────────────────

SALES_ACCOUNTS   = ['REFRESHWAVE', 'EONSPARK', 'HELLOHI', 'CUTEST CLUB']
SALES_DATA_DIR   = os.path.join(_data_dir, 'sales_data')
os.makedirs(SALES_DATA_DIR, exist_ok=True)
SALES_TTL_DAYS   = 60

REASON_MAP = {
    'order_cancelled':             'Buyer Cancelled',
    'ORC_validated with customer': 'ORC / CS Resolved',
    'shield_cancellation':         'Shield Cancel',
    'MISSHIPMENT':                 'Misshipment',
    'Attempts_Exhausted':          'Delivery Failed',
    'MISSING_ITEM':                'Missing Item',
    'DAMAGED_PRODUCT':             'Damaged Product',
    'QUALITY_ISSUE':               'Quality Issue',
    'Shipment_EOB_Ageing':         'Ageing / EOB',
    'DEFECTIVE_PRODUCT':           'Defective Product',
    'DAMAGED_SHIPMENT':            'Damaged Shipment',
    'not_serviceable':             'Not Serviceable',
    'Attempts_Exhausted':          'Delivery Failed',
}

def _sales_path(account):
    safe = re.sub(r'[^A-Za-z0-9_]', '_', account.upper())
    return os.path.join(SALES_DATA_DIR, f'sales_{safe}.json')

def _load_sales_store(account):
    """Load stored rows dict {date_str: [row, ...]} for an account."""
    p = _sales_path(account)
    if not os.path.exists(p):
        return {}
    try:
        with open(p) as f:
            return json.load(f)
    except Exception:
        return {}

def _save_sales_store(account, store):
    p  = _sales_path(account)
    tmp = p + '.tmp'
    with open(tmp, 'w') as f:
        json.dump(store, f)
    os.replace(tmp, p)

def _prune_old_dates(store):
    """Remove dates older than SALES_TTL_DAYS from today."""
    cutoff = (datetime.datetime.now(tz=IST) - datetime.timedelta(days=SALES_TTL_DAYS)).strftime('%Y-%m-%d')
    return {d: rows for d, rows in store.items() if d >= cutoff}

def _extract_product(sku):
    s = re.sub(r'\s*(Pack\s*\d+\w*|PCK\d*|\d+\s*PCK)\s*$', '', sku, flags=re.IGNORECASE).strip()
    s = re.sub(r'\s+(ES|RW|CC|HH)\s*$', '', s, flags=re.IGNORECASE).strip()
    return s or sku

def _df_to_store_rows(df):
    """Convert a filtered DataFrame to a {date_str: [row,...]} dict."""
    import pandas as pd
    store = defaultdict(list)
    for _, r in df.iterrows():
        d = str(r['date_str'])
        store[d].append({
            'sku':        r.get('sku_clean',''),
            'product':    r.get('product',''),
            'qty':        int(r.get('quantity', 1)),
            'status':     str(r.get('order_item_status','')),
            'ret_reason': str(r.get('return_reason','')) if str(r.get('return_reason','')) != 'nan' else '',
            'ret_sub':    str(r.get('return_sub_reason','')) if str(r.get('return_sub_reason','')) != 'nan' else '',
            'disp_breach':str(r.get('dispatch_sla_breached','')),
            'dlv_breach': str(r.get('delivery_sla_breached','')),
            'revenue':    float(r['selling_price']) * int(r.get('quantity', 1))
                          if 'selling_price' in df.columns and not pd.isna(r.get('selling_price'))
                          else 0.0,
        })
    return dict(store)


# ─────────────────────────────────────────────
# FLIPKART API — SALES SYNC HELPERS
# ─────────────────────────────────────────────

def _fk_compute_sla_breach(dispatch_by, dispatched_on, deliver_by, delivered_on):
    """Derive SLA breach flags from raw date strings. Returns ('Y'|'N'|'', 'Y'|'N'|'')"""
    def _parse(s):
        if not s: return None
        for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
            try: return datetime.datetime.strptime(str(s)[:19], fmt)
            except Exception: pass
        return None
    db  = _parse(dispatch_by);   don  = _parse(dispatched_on)
    dlb = _parse(deliver_by);    dlon = _parse(delivered_on)
    disp_breach = ('Y' if don > db else 'N') if db and don else ''
    dlv_breach  = ('Y' if dlon > dlb else 'N') if dlb and dlon else ''
    return disp_breach, dlv_breach


def _fk_item_to_store_row(item, product_resolver):
    """Convert one orderItem dict from Flipkart API into our store row format."""
    sku_raw   = item.get('sku', '')
    sku_clean = re.sub(r'^"""SKU:|"""$', '', sku_raw).strip().strip('"')
    product   = product_resolver(sku_clean)
    qty       = int(item.get('quantity', 1))
    status    = item.get('status', '')
    price_comp = item.get('priceComponents', {}) or {}
    unit_price = float(price_comp.get('sellingPrice') or price_comp.get('customerPrice') or 0)
    revenue    = unit_price * qty
    ret_reason = (item.get('cancellationReason', '') or '').lower()
    ret_sub    = item.get('cancellationSubReason', '') or ''
    disp_breach, dlv_breach = _fk_compute_sla_breach(
        item.get('dispatchByDate'), item.get('dispatchedDate'),
        item.get('deliverByDate'),  item.get('deliveryDate'),
    )
    order_date = item.get('orderDate', '')
    date_str   = str(order_date)[:10] if order_date else ''
    return date_str, {
        'order_item_id': item.get('orderItemId', ''),   # for deduplication on merge
        'sku': sku_clean, 'product': product, 'qty': qty, 'status': status,
        'ret_reason': ret_reason, 'ret_sub': ret_sub,
        'disp_breach': disp_breach, 'dlv_breach': dlv_breach, 'revenue': revenue,
    }

def _is_genuine_return(row, valid_reasons):
    """A RETURNED row is a genuine return only if its reason is in valid_reasons.
    Otherwise it's treated as a cancellation (e.g. order_cancelled, shield_cancellation)."""
    if row.get('status') != 'RETURNED':
        return False
    reason = (row.get('ret_reason') or '').strip().lower()
    return reason in valid_reasons

def _compute_analytics(store):
    """Compute KPIs + chart data from a {date_str: [row,...]} store dict."""
    ACTIVE    = {'DELIVERED','READY_TO_SHIP','APPROVED','APPROVAL_HOLD'}
    RETURNED  = {'RETURNED'}
    CANCELLED = {'CANCELLED','RETURN_REQUESTED'}
    # Only these return reasons count as genuine returns; all others → cancellation
    VALID_RETURN_REASONS = {
        'orc_validated with customer',
        'quality_issue',
        'missing_item',
        'customer rejection',
    }

    all_rows = [r for rows in store.values() for r in rows]
    if not all_rows:
        return None

    total_units    = sum(r['qty'] for r in all_rows)
    delivered      = sum(r['qty'] for r in all_rows if r['status'] == 'DELIVERED')
    returned_cnt   = sum(r['qty'] for r in all_rows if _is_genuine_return(r, VALID_RETURN_REASONS))
    cancelled_cnt  = sum(r['qty'] for r in all_rows if r['status'] in CANCELLED
                         or (r['status'] in RETURNED and not _is_genuine_return(r, VALID_RETURN_REASONS)))
    dispatch_breach = sum(1 for r in all_rows if r['disp_breach'] == 'Y')
    delivery_breach = sum(1 for r in all_rows if r['dlv_breach']  == 'Y')
    dispatched_tot  = sum(1 for r in all_rows if r['disp_breach'] in ('Y','N'))
    delivered_tot   = sum(1 for r in all_rows if r['dlv_breach']  in ('Y','N'))
    # Revenue — sum sellingPrice × qty for active (non-cancelled, non-returned) orders
    total_revenue   = sum(r.get('revenue', 0) for r in all_rows if r['status'] in ACTIVE)

    all_dates  = sorted(store.keys())
    date_from  = all_dates[0]  if all_dates else ''
    date_to    = all_dates[-1] if all_dates else ''

    kpis = {
        'total_orders':        len(all_rows),
        'total_units':         total_units,
        'total_revenue':       round(total_revenue, 2),
        'delivered':           delivered,
        'returned':            returned_cnt,
        'cancelled':           cancelled_cnt,
        'delivered_pct':       round(delivered     / total_units * 100, 1) if total_units else 0,
        'returned_pct':        round(returned_cnt  / total_units * 100, 1) if total_units else 0,
        'cancelled_pct':       round(cancelled_cnt / total_units * 100, 1) if total_units else 0,
        'dispatch_breach_pct': round(dispatch_breach / dispatched_tot * 100, 1) if dispatched_tot else 0,
        'delivery_breach_pct': round(delivery_breach / delivered_tot  * 100, 1) if delivered_tot  else 0,
        'date_from': date_from,
        'date_to':   date_to,
    }

    # Product trend + returns/cancels — single pass for all per-product daily data
    prod_day          = defaultdict(lambda: defaultdict(int))
    prod_total        = defaultdict(int)
    prod_revenue      = defaultdict(float)          # product → total revenue
    prod_rev_daily    = defaultdict(lambda: defaultdict(float))  # product → date → revenue
    prod_ret_daily    = defaultdict(lambda: defaultdict(int))
    prod_cancel_daily = defaultdict(lambda: defaultdict(int))
    prod_orders_daily = defaultdict(lambda: defaultdict(int))
    daily_revenue     = defaultdict(float)          # date → revenue
    for d, rows in store.items():
        for r in rows:
            prod_orders_daily[r['product']][d] += r['qty']
            if r['status'] in ACTIVE:
                prod_day[r['product']][d]   += r['qty']
                prod_total[r['product']]    += r['qty']
                rev = r.get('revenue', 0)
                prod_revenue[r['product']]         += rev
                prod_rev_daily[r['product']][d]    += rev
                daily_revenue[d]                   += rev
            if _is_genuine_return(r, VALID_RETURN_REASONS):
                prod_ret_daily[r['product']][d] += r['qty']
            elif r['status'] in CANCELLED or r['status'] in RETURNED:
                prod_cancel_daily[r['product']][d] += r['qty']

    top15 = [p for p, _ in sorted(prod_total.items(), key=lambda x: -x[1])[:15]]
    trend_series = []
    for prod in top15:
        trend_series.append({
            'name':        prod,
            'data':        {d: prod_day[prod].get(d, 0) for d in all_dates},
            'data_ret':    {d: prod_ret_daily[prod].get(d, 0) for d in all_dates},
            'data_cancel': {d: prod_cancel_daily[prod].get(d, 0) for d in all_dates},
            'revenue':     round(prod_revenue.get(prod, 0), 2),
            'rev_daily':   {d: round(prod_rev_daily[prod].get(d, 0), 2) for d in all_dates},
        })

    # Returns chart
    reason_counts = defaultdict(int)
    for r in all_rows:
        if r['status'] in (RETURNED | CANCELLED):
            label = REASON_MAP.get(r['ret_reason'], r['ret_reason']) if r['ret_reason'] else 'Unknown'
            reason_counts[label] += 1
    returns_chart = [
        {'reason': k, 'count': v}
        for k, v in sorted(reason_counts.items(), key=lambda x: -x[1])[:12]
    ]

    # ── Returns by product — aggregates from already-computed daily data ─────

    # Aggregate totals (full range) for sorting/ranking
    prod_returns = {p: sum(v.values()) for p, v in prod_ret_daily.items()}
    prod_orders  = {p: sum(v.values()) for p, v in prod_orders_daily.items()}
    prod_cancels_total = {p: sum(v.values()) for p, v in prod_cancel_daily.items()}

    top_ret_prods = sorted(prod_returns.items(), key=lambda x: -x[1])[:15]
    returns_by_product = [
        {
            'product':      p,
            'returns':      cnt,
            'total':        prod_orders.get(p, 0),
            'rate':         round(cnt / prod_orders[p] * 100, 1) if prod_orders.get(p) else 0,
            'daily_ret':    dict(prod_ret_daily[p]),
            'daily_orders': dict(prod_orders_daily[p]),
        }
        for p, cnt in top_ret_prods
    ]

    # ── Return reason drill-down: reason -> {sub_reason: count} ─────────────
    reason_drill = defaultdict(lambda: defaultdict(int))
    for r in all_rows:
        if _is_genuine_return(r, VALID_RETURN_REASONS):
            label = REASON_MAP.get(r['ret_reason'], r['ret_reason']) if r['ret_reason'] else 'Unknown'
            sub   = r.get('ret_sub', '') or r.get('ret_reason', '') or 'Unknown'
            reason_drill[label][sub] += 1
    returns_drill = {
        reason: [{'sub': s, 'count': c} for s, c in sorted(subs.items(), key=lambda x: -x[1])]
        for reason, subs in reason_drill.items()
    }

    # ── Daily return/cancel data for client-side date filtering ─────────────
    daily_returns  = defaultdict(int)   # date -> return qty
    daily_orders   = defaultdict(int)   # date -> total qty
    daily_cancels  = defaultdict(int)   # date -> cancel qty
    for d, rows in store.items():
        for r in rows:
            daily_orders[d] += r['qty']
            if _is_genuine_return(r, VALID_RETURN_REASONS):
                daily_returns[d] += r['qty']
            elif r['status'] in CANCELLED or r['status'] in RETURNED:
                daily_cancels[d] += r['qty']

    # ── Weekly return trend ───────────────────────────────────────────────────
    import datetime as _dt
    week_returns = defaultdict(int)
    week_orders  = defaultdict(int)
    for d in all_dates:
        try:
            dt = _dt.datetime.strptime(d, '%Y-%m-%d')
            wk = dt.strftime('%G-W%V')   # ISO week: 2026-W12
        except Exception:
            continue
        week_orders[wk]  += daily_orders.get(d, 0)
        week_returns[wk] += daily_returns.get(d, 0)
    all_weeks = sorted(set(list(week_orders.keys()) + list(week_returns.keys())))
    return_trend = [
        {
            'week':    w,
            'returns': week_returns.get(w, 0),
            'orders':  week_orders.get(w, 0),
            'rate':    round(week_returns.get(w, 0) / week_orders.get(w, 1) * 100, 1),
        }
        for w in all_weeks
    ]

    # ── Cancellations by product (top 15) ─────────────────────────────────────
    top_cancel_prods = sorted(prod_cancels_total.items(), key=lambda x: -x[1])[:15]
    cancels_by_product = [
        {
            'product':         p,
            'cancels':         cnt,
            'total':           prod_orders.get(p, 0),
            'rate':            round(cnt / prod_orders[p] * 100, 1) if prod_orders.get(p) else 0,
            'daily_cancel':    dict(prod_cancel_daily[p]),
            'daily_orders':    dict(prod_orders_daily[p]),
        }
        for p, cnt in top_cancel_prods
    ]

    # ── Per-date return/cancel totals for client-side date filter ─────────────
    daily_stats = [
        {
            'date':    d,
            'returns': daily_returns.get(d, 0),
            'cancels': daily_cancels.get(d, 0),
            'orders':  daily_orders.get(d, 0),
            'revenue': round(daily_revenue.get(d, 0), 2),
        }
        for d in all_dates
    ]

    # ── SKU breakdown per product (active orders only) ──────────────────────
    # Structure: { product: { sku: {total, daily: {date: qty}, revenue} } }
    prod_sku_daily   = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    prod_sku_total   = defaultdict(lambda: defaultdict(int))
    prod_sku_revenue = defaultdict(lambda: defaultdict(float))
    for d, rows in store.items():
        for r in rows:
            if r['status'] in ACTIVE:
                prod_sku_daily[r['product']][r['sku']][d]   += r['qty']
                prod_sku_total[r['product']][r['sku']]      += r['qty']
                prod_sku_revenue[r['product']][r['sku']]    += r.get('revenue', 0)

    # For each product: store top 50 SKUs (table needs all; donut/line cap at 7 on frontend)
    SKU_STORE_LIMIT = 50
    sku_by_product = {}
    for prod, sku_totals in prod_sku_total.items():
        top_skus_list = sorted(sku_totals.items(), key=lambda x: -x[1])[:SKU_STORE_LIMIT]
        top_skus      = [s for s, _ in top_skus_list]
        others_daily  = defaultdict(int)
        for sku, day_map in prod_sku_daily[prod].items():
            if sku not in top_skus:
                for d, q in day_map.items():
                    others_daily[d] += q
        series = []
        for sku in top_skus:
            series.append({
                'sku':     sku,
                'total':   prod_sku_total[prod][sku],
                'revenue': round(prod_sku_revenue[prod].get(sku, 0), 2),
                'daily':   {d: prod_sku_daily[prod][sku].get(d, 0) for d in all_dates},
            })
        if others_daily:
            series.append({
                'sku':     'Others',
                'total':   sum(others_daily.values()),
                'revenue': 0.0,
                'daily':   {d: others_daily.get(d, 0) for d in all_dates},
            })
        sku_by_product[prod] = series

    return {
        'kpis':               kpis,
        'trend_dates':        all_dates,
        'trend_series':       trend_series,
        'daily_revenue':      {d: round(daily_revenue.get(d, 0), 2) for d in all_dates},
        'returns_chart':      returns_chart,
        'returns_by_product': returns_by_product,
        'returns_drill':      returns_drill,
        'return_trend':       return_trend,
        'cancels_by_product': cancels_by_product,
        'daily_stats':        daily_stats,
        'sku_by_product':     sku_by_product,
    }



def _fk_sync_sales(account, full_resync=False):
    """
    Fetch recent shipments from Flipkart API and merge into the persistent sales store.
    If full_resync=True, wipes the store first and fetches the full 60-day window.
    Otherwise incremental — only fetches dates after the latest date already stored (minus 2-day buffer).
    """
    import requests as _req

    token   = fk_get_token(account)
    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    base_url = f'{FK_API_BASE}/sellers/v3/shipments/filter/'

    now_ist   = datetime.datetime.now(tz=IST)
    cutoff    = (now_ist - datetime.timedelta(days=SALES_TTL_DAYS)).strftime('%Y-%m-%d')

    store     = _load_sales_store(account)
    store     = _prune_old_dates(store)
    if full_resync:
        # Full re-sync: fetch the entire 60-day window but KEEP existing rows —
        # upsert will overwrite stale API rows while preserving XLSX-uploaded history
        fetch_from = cutoff
    else:
        existing_dates = [k for k in store if not k.startswith('__')]
        if existing_dates:
            latest     = max(existing_dates)
            fetch_from = (datetime.datetime.strptime(latest, '%Y-%m-%d')
                          - datetime.timedelta(days=2)).strftime('%Y-%m-%d')
        else:
            fetch_from = cutoff
    fetch_to = now_ist.strftime('%Y-%m-%d')

    # Build product resolver
    sku_to_product = {}
    if os.path.exists(MASTER_SKU_PATH):
        try:
            with open(MASTER_SKU_PATH, 'rb') as f:
                master = load_sku_master(f.read())
            acct_master = get_account_master(master, account)
            for sku, info in acct_master.items():
                if info.get('product'):
                    sku_to_product[sku] = info['product']
        except Exception as me:
            print(f'[FKSync] Master SKU load failed: {me}')

    def resolve_product(sku_clean):
        def norm(s): return s.strip().title()
        if sku_clean in sku_to_product: return norm(sku_to_product[sku_clean])
        stripped = re.sub(r'\s*(Pack\s*\d+\w*|PCK\d*|\d+\s*PCK)\s*$', '', sku_clean, flags=re.IGNORECASE).strip()
        if stripped != sku_clean and stripped in sku_to_product: return norm(sku_to_product[stripped])
        for ms_key, prod_name in sku_to_product.items():
            ms = re.sub(r'\s*(Pack\s*\d+\w*|PCK\d*|\d+\s*PCK)\s*$', '', ms_key, flags=re.IGNORECASE).strip()
            if ms and (ms == stripped or ms == sku_clean): return norm(prod_name)
        return norm(stripped) if stripped else sku_clean

    new_rows      = defaultdict(list)
    total_fetched = 0

    # --- preDispatch: APPROVED, READY_TO_DISPATCH, PACKED (active unfulfilled orders) ---
    payload = {
        'filter': {
            'type': 'preDispatch',
            'states': ['APPROVED', 'READY_TO_DISPATCH', 'PACKED', 'PACKING_IN_PROGRESS'],
            'orderDate': {
                'from': f'{fetch_from}T00:00:00+05:30',
                'to':   f'{fetch_to}T23:59:59+05:30',
            },
        },
        'pagination': {'pageSize': 20},
    }
    next_url = base_url
    fetched  = 0
    while next_url and fetched < 5000:
        try:
            r = (_req.post(next_url, json=payload, headers=headers, timeout=30)
                 if next_url == base_url
                 else _req.get(next_url, headers=headers, timeout=30))
            if r.status_code != 200:
                print(f'[FKSync] preDispatch {r.status_code}: {r.text[:200]}')
                break
            data = r.json()
            for shipment in data.get('shipments', []):
                for item in shipment.get('orderItems', []):
                    ds, row = _fk_item_to_store_row(item, resolve_product)
                    if ds and ds >= cutoff:
                        new_rows[ds].append(row)
                        fetched += 1
            if not data.get('hasMore') or not data.get('nextPageUrl'):
                break
            next_url = data['nextPageUrl']
        except Exception as e:
            print(f'[FKSync] preDispatch error: {e}')
            break
    total_fetched += fetched
    print(f'[FKSync] preDispatch → {fetched} items')

    # --- postDispatch: SHIPPED + DELIVERED ---
    for state_list in [['SHIPPED', 'DELIVERED']]:
        payload  = {
            'filter': {
                'type': 'postDispatch', 'states': state_list,
                'orderDate': {
                    'from': f'{fetch_from}T00:00:00+05:30',
                    'to':   f'{fetch_to}T23:59:59+05:30',
                },
            },
            'pagination': {'pageSize': 20},
        }
        next_url = base_url
        fetched  = 0
        while next_url and fetched < 5000:
            try:
                r = (_req.post(next_url, json=payload, headers=headers, timeout=30)
                     if next_url == base_url
                     else _req.get(next_url, headers=headers, timeout=30))
                if r.status_code != 200:
                    print(f'[FKSync] postDispatch {r.status_code}: {r.text[:200]}')
                    break
                data = r.json()
                for shipment in data.get('shipments', []):
                    for item in shipment.get('orderItems', []):
                        ds, row = _fk_item_to_store_row(item, resolve_product)
                        if ds and ds >= cutoff:
                            new_rows[ds].append(row)
                            fetched += 1
                if not data.get('hasMore') or not data.get('nextPageUrl'):
                    break
                next_url = data['nextPageUrl']
            except Exception as e:
                print(f'[FKSync] postDispatch error: {e}')
                break
        total_fetched += fetched
        print(f'[FKSync] postDispatch → {fetched} items')

    # --- cancelled orders (all 3 cancellation types) ---
    for ctype in ['buyerCancellation', 'sellerCancellation', 'marketplaceCancellation']:
        payload = {
            'filter': {
                'type': 'cancelled', 'states': ['CANCELLED'],
                'cancellationType': ctype,
                'cancellationDate': {
                    'from': f'{fetch_from}T00:00:00+05:30',
                    'to':   f'{fetch_to}T23:59:59+05:30',
                },
            },
            'pagination': {'pageSize': 20},
        }
        next_url = base_url
        fetched  = 0
        while next_url and fetched < 5000:
            try:
                r = (_req.post(next_url, json=payload, headers=headers, timeout=30)
                     if next_url == base_url
                     else _req.get(next_url, headers=headers, timeout=30))
                if r.status_code != 200: break
                data = r.json()
                for shipment in data.get('shipments', []):
                    for item in shipment.get('orderItems', []):
                        ds, row = _fk_item_to_store_row(item, resolve_product)
                        if ds and ds >= cutoff:
                            new_rows[ds].append(row)
                            fetched += 1
                if not data.get('hasMore') or not data.get('nextPageUrl'):
                    break
                next_url = data['nextPageUrl']
            except Exception as e:
                print(f'[FKSync] {ctype} error: {e}')
                break
        total_fetched += fetched
        print(f'[FKSync] {ctype} → {fetched} items')

    # --- returns via /v2/returns ---
    try:
        r = _req.get(f'{FK_API_BASE}/sellers/v2/returns', headers=headers, timeout=30)
        if r.status_code == 200:
            ret_fetched = 0
            for ret in r.json().get('returnItems', []):
                oi         = ret.get('orderItem', {}) or {}
                sku_raw    = oi.get('sku', '')
                sku_clean  = re.sub(r'^\"\"\"SKU:|\"\"\"$', '', sku_raw).strip().strip('"')
                product    = resolve_product(sku_clean)
                qty        = int(oi.get('quantity', 1))
                order_date = oi.get('orderDate', '')
                ds         = str(order_date)[:10] if order_date else ''
                pc         = oi.get('priceComponents', {}) or {}
                revenue    = float(pc.get('sellingPrice') or 0) * qty
                if ds and ds >= cutoff:
                    new_rows[ds].append({
                        'sku': sku_clean, 'product': product, 'qty': qty,
                        'status': 'RETURNED',
                        'ret_reason': (ret.get('returnReason') or '').lower(),
                        'ret_sub': ret.get('returnSubReason') or '',
                        'disp_breach': '', 'dlv_breach': '', 'revenue': revenue,
                    })
                    ret_fetched += 1
            total_fetched += ret_fetched
            print(f'[FKSync] returns → {ret_fetched} items')
    except Exception as e:
        print(f'[FKSync] Returns fetch error: {e}')

    # Merge new_rows into store — upsert by order_item_id so we never lose existing orders.
    # For each date: build a map of existing rows keyed by order_item_id, then overwrite
    # with fresh API rows (which have updated status), keeping any rows the API didn't return.
    for d, api_rows in new_rows.items():
        existing = store.get(d, [])
        # Index existing rows by order_item_id (rows without one are kept as-is)
        indexed = {r['order_item_id']: r for r in existing if r.get('order_item_id')}
        unkeyed = [r for r in existing if not r.get('order_item_id')]
        # Overwrite/add API rows
        for row in api_rows:
            oid = row.get('order_item_id', '')
            if oid:
                indexed[oid] = row
            else:
                unkeyed.append(row)
        store[d] = list(indexed.values()) + unkeyed

    store['__meta__'] = {
        'updated_at':   now_ist.strftime('%d %b %Y, %H:%M IST'),
        'account':      account,
        'sync_method':  'api',
        'fetch_from':   fetch_from,
        'fetch_to':     fetch_to,
    }
    _save_sales_store(account, store)

    return {
        'ok': True, 'account': account,
        'total_fetched': total_fetched,
        'fetch_from': fetch_from, 'fetch_to': fetch_to,
        'new_dates': len(new_rows),
    }

@app.route('/api/sales-upload/<account>', methods=['POST'])
def sales_upload(account):
    """
    Upload an Orders XLSX for one account. Merges into persistent store,
    newer upload wins on overlapping dates. Prunes data older than 60 days.
    """
    account = account.upper().replace('-', ' ')
    xlsx_file = request.files.get('xlsx')
    if not xlsx_file:
        return jsonify({'error': 'No file provided'}), 400
    try:
        import pandas as pd
        from io import BytesIO

        data = xlsx_file.read()
        df   = pd.read_excel(BytesIO(data), sheet_name='Orders')

        df['sku_clean']  = df['sku'].str.replace(r'"""SKU:', '', regex=True).str.replace('"""', '', regex=True).str.strip()
        df['order_date'] = pd.to_datetime(df['order_date'], errors='coerce')
        df['date_str']   = df['order_date'].dt.strftime('%Y-%m-%d')

        # ── Map SKU → product name using Master SKU file ──────────────────────
        # Falls back to regex-extracted name if master not available or SKU not found
        sku_to_product = {}
        if os.path.exists(MASTER_SKU_PATH):
            try:
                with open(MASTER_SKU_PATH, 'rb') as f:
                    master = load_sku_master(f.read())
                acct_master = get_account_master(master, account)
                for sku, info in acct_master.items():
                    if info.get('product'):
                        sku_to_product[sku] = info['product']
            except Exception as me:
                print(f'[SalesUpload] Master SKU lookup failed: {me}')

        def resolve_product(sku_clean):
            # Normalise helper — title-case for consistent cross-account grouping
            def norm(s): return s.strip().title()
            # 1. Exact match
            if sku_clean in sku_to_product:
                return norm(sku_to_product[sku_clean])
            # 2. Strip pack suffix from orders SKU and try again
            #    e.g. "Zippd Natural Detox Foot Pad 3PCK" -> "Zippd Natural Detox Foot Pad"
            stripped = re.sub(r'\s*(Pack\s*\d+\w*|PCK\d*|\d+\s*PCK)\s*$', '', sku_clean, flags=re.IGNORECASE).strip()
            if stripped != sku_clean and stripped in sku_to_product:
                return norm(sku_to_product[stripped])
            # 3. Strip pack suffix from master SKU keys and compare
            for master_sku, prod_name in sku_to_product.items():
                ms = re.sub(r'\s*(Pack\s*\d+\w*|PCK\d*|\d+\s*PCK)\s*$', '', master_sku, flags=re.IGNORECASE).strip()
                if ms and (ms == stripped or ms == sku_clean):
                    return norm(prod_name)
            # 4. Fallback: return regex-stripped SKU (no master match found)
            return norm(stripped) if stripped else sku_clean

        df['product'] = df['sku_clean'].apply(resolve_product)

        # Filter to last 60 days
        cutoff = (datetime.datetime.now(tz=IST) - datetime.timedelta(days=SALES_TTL_DAYS)).strftime('%Y-%m-%d')
        df = df[df['date_str'] >= cutoff].copy()

        # Convert to store rows
        new_rows = _df_to_store_rows(df)

        # Load existing, prune, then overlay new dates (newer wins)
        store = _load_sales_store(account)
        store = _prune_old_dates(store)
        # Explicitly remove existing entries for dates in the new upload
        # so corrected product names take effect immediately on re-upload
        for d in new_rows:
            store.pop(d, None)
        store.update(new_rows)

        # Record upload timestamp
        store['__meta__'] = {
            'updated_at': datetime.datetime.now(tz=IST).strftime('%d %b %Y, %H:%M IST'),
            'account':    account,
        }
        _save_sales_store(account, store)

        # Return computed analytics immediately
        clean_store = {k: v for k, v in store.items() if not k.startswith('__')}
        result = _compute_analytics(clean_store)
        if result:
            result['updated_at'] = store['__meta__']['updated_at']
        return jsonify(result or {'error': 'No data after filtering'})

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/sales-debug/<account>', methods=['GET'])
def sales_debug(account):
    """Debug: show raw stored data summary for an account."""
    account = account.upper().replace('-', ' ')
    store = _load_sales_store(account)
    all_rows = [r for k, v in store.items() if not k.startswith('__') for r in v]
    RETURNED = {'RETURNED','RETURN_REQUESTED'}

    # Per-product returns
    product_returns = {}
    for r in all_rows:
        p = r.get('product','?')
        if r.get('status') in RETURNED:
            product_returns[p] = product_returns.get(p, 0) + r.get('qty', 1)
    top = sorted(product_returns.items(), key=lambda x: -x[1])[:20]

    # SKU-level breakdown for sp4 toothpaste diagnosis
    sku_returns = {}
    for r in all_rows:
        if r.get('status') in RETURNED:
            key = (r.get('sku','?'), r.get('product','?'))
            sku_returns[key] = sku_returns.get(key, 0) + r.get('qty', 1)
    # Filter to toothpaste-related
    tooth_skus = {k: v for k, v in sku_returns.items()
                  if any(x in k[0].lower() or x in k[1].lower()
                         for x in ['tooth','sp4','sp 4','herbheal','manicure'])}
    tooth_list = sorted(tooth_skus.items(), key=lambda x: -x[1])

    return jsonify({
        'account': account,
        'total_rows': len(all_rows),
        'date_range': [min((k for k in store if not k.startswith('__')), default=''), max((k for k in store if not k.startswith('__')), default='')],
        'unique_dates': len([k for k in store if not k.startswith('__')]),
        'top_returns_by_product': [{'product': p, 'returns': c} for p, c in top],
        'tooth_sku_breakdown': [{'sku': k[0], 'product': k[1], 'returns': v} for k, v in tooth_list],
        'meta': store.get('__meta__', {}),
    })

@app.route('/api/sales-clear/<account>', methods=['POST'])
def sales_clear(account):
    """Delete stored sales data for an account so it can be re-uploaded cleanly."""
    account = account.upper().replace('-', ' ')
    p = _sales_path(account)
    if os.path.exists(p):
        os.remove(p)
    return jsonify({'ok': True, 'account': account})


@app.route('/api/sales-data/<account>', methods=['GET'])
def sales_data(account):
    """Return stored analytics for one account (no upload needed)."""
    account = account.upper().replace('-', ' ')
    if account == 'CONSOLIDATED':
        # Merge all accounts
        merged = {}
        for acc in SALES_ACCOUNTS:
            store = _load_sales_store(acc)
            store = _prune_old_dates(store)
            for d, rows in store.items():
                if d.startswith('__'):
                    continue
                if d not in merged:
                    merged[d] = []
                merged[d].extend(rows)
        result = _compute_analytics(merged)
        if not result:
            return jsonify({'empty': True})
        # Collect per-account last-updated
        updates = {}
        for acc in SALES_ACCOUNTS:
            s = _load_sales_store(acc)
            meta = s.get('__meta__', {})
            if meta.get('updated_at'):
                updates[acc] = meta['updated_at']
        result['account_updates'] = updates
        return jsonify(result)
    else:
        store = _load_sales_store(account)
        store = _prune_old_dates(store)
        meta  = store.pop('__meta__', {})
        clean = {k: v for k, v in store.items() if not k.startswith('__')}
        if not clean:
            return jsonify({'empty': True})
        result = _compute_analytics(clean)
        if result and meta.get('updated_at'):
            result['updated_at'] = meta['updated_at']
        return jsonify(result or {'empty': True})


@app.route('/api/sales-sync/<account>', methods=['POST'])
def sales_sync(account):
    """
    Trigger a Flipkart API sync for the given account.
    Only works for accounts with FK credentials configured.
    """
    account = account.upper().replace('-', ' ')
    if account not in SALES_ACCOUNTS:
        return jsonify({'error': f'Unknown account: {account}'}), 400
    app_id, _ = _fk_credentials(account)
    if not app_id:
        return jsonify({'error': f'No Flipkart API credentials configured for {account}. '
                                  f'Set {_fk_env_key(account)}_APP_ID and _APP_SECRET in Railway.'}), 400
    try:
        body = request.get_json() or {}
        full_resync = bool(body.get('full_resync', False))
        result = _fk_sync_sales(account, full_resync=full_resync)
        # Return fresh analytics immediately after sync
        store  = _load_sales_store(account)
        store  = _prune_old_dates(store)
        meta   = store.pop('__meta__', {})
        clean  = {k: v for k, v in store.items() if not k.startswith('__')}
        analytics = _compute_analytics(clean)
        if analytics and meta.get('updated_at'):
            analytics['updated_at'] = meta['updated_at']
        return jsonify({
            'sync':      result,
            'analytics': analytics or {},
        })
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/sales-sync-status', methods=['GET'])
def sales_sync_status():
    """Return sync status for all accounts — which have API credentials, last sync time."""
    results = {}
    for account in SALES_ACCOUNTS:
        app_id, _ = _fk_credentials(account)
        store = _load_sales_store(account)
        meta  = store.get('__meta__', {})
        results[account] = {
            'api_configured': bool(app_id),
            'sync_method':    meta.get('sync_method', 'xlsx'),
            'updated_at':     meta.get('updated_at', 'never'),
            'fetch_from':     meta.get('fetch_from', ''),
            'fetch_to':       meta.get('fetch_to', ''),
        }
    return jsonify(results)


# ─────────────────────────────────────────────
# ADS ANALYTICS  helpers
# ─────────────────────────────────────────────

ADS_DATA_DIR  = os.path.join(_data_dir, 'ads_data')
os.makedirs(ADS_DATA_DIR, exist_ok=True)
ADS_TTL_DAYS  = 60

# How many metadata rows to skip per report type
ADS_SKIP_ROWS = {
    'campaignOrder':   2,
    'consolidated':    2,
    'consolidatedFSN': 2,
    'keyword':         2,
    'pla':             4,
    'placement':       2,
    'searchTerm':      2,
}

def _ads_path(account):
    safe = re.sub(r'[^A-Za-z0-9_]', '_', account.upper())
    return os.path.join(ADS_DATA_DIR, f'ads_{safe}.json')

def _load_ads_store(account):
    p = _ads_path(account)
    if not os.path.exists(p):
        return {}
    try:
        with open(p) as f:
            return json.load(f)
    except Exception:
        return {}

def _save_ads_store(account, store):
    p   = _ads_path(account)
    tmp = p + '.tmp'
    with open(tmp, 'w') as f:
        json.dump(store, f)
    os.replace(tmp, p)

def _prune_ads_store(store):
    """Remove date-keyed entries older than ADS_TTL_DAYS."""
    cutoff = (datetime.datetime.now(tz=IST) - datetime.timedelta(days=ADS_TTL_DAYS)).strftime('%Y-%m-%d')
    pruned = {}
    for k, v in store.items():
        # keep meta keys and non-date keys always; prune only date-shaped keys
        if re.match(r'^\d{4}-\d{2}-\d{2}$', k):
            if k >= cutoff:
                pruned[k] = v
        else:
            pruned[k] = v
    return pruned

def _parse_csv_bytes(data_bytes, report_type):
    """Parse a CSV bytes object into list-of-dicts, applying correct skip rows."""
    import pandas as pd
    import math
    from io import BytesIO
    skip = ADS_SKIP_ROWS.get(report_type, 2)
    df   = pd.read_csv(BytesIO(data_bytes), skiprows=skip)
    df.columns = [c.strip() for c in df.columns]
    rows = df.to_dict(orient='records')
    # Replace NaN/Inf with None so json.dumps produces valid JSON
    clean = []
    for row in rows:
        clean.append({
            k: (None if isinstance(v, float) and (math.isnan(v) or math.isinf(v)) else v)
            for k, v in row.items()
        })
    return clean

def _merge_ads_rows(existing_rows, new_rows):
    """Merge two lists of row dicts — new rows appended, no dedup needed (date-bucketed)."""
    return new_rows   # for date-bucketed store, new fully replaces

@app.route('/api/ads-upload/<account>', methods=['POST'])
def ads_upload(account):
    """
    Receive up to 7 CSVs for one ads account. Parse server-side, merge into
    persistent JSON store (60-day TTL, newer upload wins per date bucket).
    """
    account = account.upper().replace('-', ' ')
    try:
        import pandas as pd

        # Load + prune existing store
        store = _load_ads_store(account)
        store = _prune_ads_store(store)

        # Collect uploaded files
        file_keys = ['campaignOrder', 'consolidated', 'consolidatedFSN',
                     'keyword', 'pla', 'placement', 'searchTerm']
        parsed = {}
        date_from, date_to = None, None

        for key in file_keys:
            f = request.files.get(key)
            if not f:
                continue
            raw = f.read()

            # Extract date range from first two rows
            import io as _io
            text_lines = raw.decode('utf-8-sig', errors='replace').splitlines()
            if len(text_lines) > 0 and 'Start Time' in text_lines[0]:
                parts = text_lines[0].split(',')
                if len(parts) > 1:
                    date_from = parts[1].strip()[:10]   # YYYY-MM-DD
            if len(text_lines) > 1 and 'End Time' in text_lines[1]:
                parts = text_lines[1].split(',')
                if len(parts) > 1:
                    date_to = parts[1].strip()[:10]

            try:
                rows = _parse_csv_bytes(raw, key)
                parsed[key] = rows
            except Exception as pe:
                print(f'[AdsUpload] Failed to parse {key}: {pe}')
                continue

        if not parsed:
            return jsonify({'error': 'No valid files uploaded'}), 400

        # Cap large reports to keep stored JSON manageable
        ROW_CAPS = {
            'searchTerm':    2000,   # 54k rows is too large; top 2000 by views is plenty
            'campaignOrder': 5000,
            'keyword':       3000,
            'placement':     2000,
        }
        for key, cap in ROW_CAPS.items():
            if key in parsed and len(parsed[key]) > cap:
                # Sort by Views desc if available, else just truncate
                rows = parsed[key]
                try:
                    rows = sorted(rows, key=lambda r: float(r.get('Views', 0) or 0), reverse=True)
                except Exception:
                    pass
                parsed[key] = rows[:cap]

        # Auto-correct consolidated/consolidatedFSN swap:
        # The correct 'consolidated' key must have 'Ad Spend' (campaign-level spend data).
        # If the user uploaded them in the wrong slots, fix it here at storage time.
        def _has_spend(rows):
            return bool(rows) and 'Ad Spend' in (rows[0] if rows else {})

        c     = parsed.get('consolidated', [])
        c_fsn = parsed.get('consolidatedFSN', [])
        if c and c_fsn:
            if not _has_spend(c) and _has_spend(c_fsn):
                # Swap: consolidatedFSN has spend, consolidated has order data
                parsed['consolidated'], parsed['consolidatedFSN'] = c_fsn, c
        elif c and not _has_spend(c) and not c_fsn:
            # Only consolidated uploaded but it's the wrong one — move it
            parsed['consolidatedFSN'] = c
            del parsed['consolidated']
        elif c_fsn and _has_spend(c_fsn) and not c:
            # Only consolidatedFSN uploaded and it has spend — promote it
            parsed['consolidated'] = c_fsn
            del parsed['consolidatedFSN']

        # Store under a composite key: date range + each report type
        # Use date_from as the bucket key (or 'undated' fallback)
        bucket_key = date_from or 'undated'

        # Merge: for same bucket, new upload wins entirely
        if bucket_key not in store:
            store[bucket_key] = {}
        for key, rows in parsed.items():
            store[bucket_key][key] = rows

        # Also keep track of overall date range in meta
        store['__meta__'] = {
            'updated_at': datetime.datetime.now(tz=IST).strftime('%d %b %Y, %H:%M IST'),
            'account':    account,
            'date_from':  date_from,
            'date_to':    date_to,
        }

        _save_ads_store(account, store)

        # Return merged data across all buckets
        result = _build_ads_response(store)
        return jsonify(result)

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/ads-data/<account>', methods=['GET'])
def ads_data_get(account):
    """Return stored ads analytics for one account, or CONSOLIDATED."""
    account = account.upper().replace('-', ' ')

    if account == 'CONSOLIDATED':
        merged_store = {'__meta__': {'account': 'CONSOLIDATED'}}
        any_data = False
        for acc in SALES_ACCOUNTS:   # reuse same account list
            s = _load_ads_store(acc)
            s = _prune_ads_store(s)
            for bucket, reports in s.items():
                if bucket.startswith('__'):
                    continue
                any_data = True
                if bucket not in merged_store:
                    merged_store[bucket] = {}
                for report_key, rows in reports.items():
                    if report_key not in merged_store[bucket]:
                        merged_store[bucket][report_key] = []
                    merged_store[bucket][report_key].extend(rows)
        if not any_data:
            return jsonify({'empty': True})
        result = _build_ads_response(merged_store)
        # Add per-account update times
        updates = {}
        for acc in SALES_ACCOUNTS:
            s = _load_ads_store(acc)
            meta = s.get('__meta__', {})
            if meta.get('updated_at'):
                updates[acc] = meta['updated_at']
        result['account_updates'] = updates
        return jsonify(result)
    else:
        store = _load_ads_store(account)
        store = _prune_ads_store(store)
        if not any(not k.startswith('__') for k in store):
            return jsonify({'empty': True})
        result = _build_ads_response(store)
        return jsonify(result)




@app.route('/api/ads-colcheck/<account>', methods=['GET'])
def ads_colcheck(account):
    """Debug: show column names in stored campaignOrder for an account."""
    account = account.upper().replace('-', ' ')
    store = _load_ads_store(account)
    store = _prune_ads_store(store)
    result = _build_ads_response(store)
    data = result.get('data', {})
    co = data.get('campaignOrder', [])
    return jsonify({
        'account': account,
        'campaignOrder_rows': len(co),
        'campaignOrder_columns': list(co[0].keys()) if co else [],
        'campaignOrder_sample': co[0] if co else {},
        'available_report_types': list(data.keys()),
    })

@app.route('/api/ads-status/<account>', methods=['GET'])
def ads_status(account):
    """Quick status check — returns meta + bucket keys without full data."""
    account = account.upper().replace('-', ' ')
    store = _load_ads_store(account)
    if not store:
        return jsonify({'exists': False})
    meta = store.get('__meta__', {})
    buckets = {k: list(v.keys()) if isinstance(v, dict) else type(v).__name__
               for k, v in store.items() if not k.startswith('__')}
    return jsonify({
        'exists': True,
        'meta': meta,
        'buckets': buckets,
        'file_path': _ads_path(account),
    })

@app.route('/api/ads-debug/<account>', methods=['GET'])
def ads_debug(account):
    """Temporary debug endpoint — shows what's stored for an account."""
    account = account.upper().replace('-', ' ')
    store = _load_ads_store(account)
    result = {}
    for bucket, reports in store.items():
        if bucket.startswith('__'):
            result[bucket] = reports
            continue
        if not isinstance(reports, dict):
            continue
        result[bucket] = {}
        for key, rows in reports.items():
            sample = rows[0] if rows else {}
            result[bucket][key] = {
                'row_count': len(rows),
                'columns': list(sample.keys()) if sample else [],
                'sample_row': {k: str(v)[:80] for k, v in list(sample.items())[:6]} if sample else {}
            }
    return jsonify(result)

@app.route('/api/ads-clear/<account>', methods=['POST'])
def ads_clear(account):
    """Delete stored ads data for an account."""
    account = account.upper().replace('-', ' ')
    p = _ads_path(account)
    if os.path.exists(p):
        os.remove(p)
    return jsonify({'ok': True})


def _build_ads_response(store):
    """Merge all date buckets in store into a flat data dict for the frontend."""
    meta = store.get('__meta__', {})

    # Merge rows across all date buckets
    merged = {}   # report_key -> [rows]
    for bucket, reports in store.items():
        if bucket.startswith('__'):
            continue
        if not isinstance(reports, dict):
            continue
        for key, rows in reports.items():
            if key not in merged:
                merged[key] = []
            merged[key].extend(rows)

    # Auto-correct common swap: if 'consolidated' lacks 'Ad Spend' but 'consolidatedFSN'
    # has it, swap them so the frontend KPI block always reads spend from 'consolidated'
    def _has_spend(rows):
        return bool(rows) and 'Ad Spend' in (rows[0] if rows else {})

    consol     = merged.get('consolidated', [])
    consol_fsn = merged.get('consolidatedFSN', [])
    if not _has_spend(consol) and _has_spend(consol_fsn):
        merged['consolidated'], merged['consolidatedFSN'] = consol_fsn, consol

    return {
        'data':       merged,
        'updated_at': meta.get('updated_at', ''),
        'date_from':  meta.get('date_from', ''),
        'date_to':    meta.get('date_to', ''),
    }

# ─────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────

@app.route('/')
def index():
    with open(os.path.join(os.path.dirname(__file__), 'templates', 'index.html'), 'r') as f:
        return Response(f.read(), mimetype='text/html')

@app.route('/api/master-sku-map', methods=['GET'])
def master_sku_map():
    """Return SKU -> ProductName mapping from master SKU file."""
    if not os.path.exists(MASTER_SKU_PATH):
        return jsonify({'error': 'Master SKU not uploaded'}), 404
    try:
        import pandas as pd
        df = pd.read_excel(MASTER_SKU_PATH)
        mapping = {}
        for _, row in df.iterrows():
            sku = str(row.get('SKU', '')).strip()
            p1  = str(row.get('ProductName 1', '')).strip() if pd.notna(row.get('ProductName 1')) else ''
            if sku and p1 and sku != 'nan':
                mapping[sku] = p1
        return jsonify({'mapping': mapping})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/master-status')
def master_status():
    """Check if a master SKU file is stored on the server."""
    exists = os.path.exists(MASTER_SKU_PATH)
    mtime = None
    if exists:
        import datetime
        mtime = datetime.datetime.fromtimestamp(os.path.getmtime(MASTER_SKU_PATH), tz=IST).strftime('%d %b %Y, %H:%M IST')
    return jsonify({'exists': exists, 'updated': mtime})

@app.route('/api/upload-master', methods=['POST'])
def upload_master():
    """Save a new master SKU file, overwriting any existing one."""
    sku_file = request.files.get('sku_csv')
    if not sku_file:
        return jsonify({'error': 'No file provided'}), 400
    sku_file.save(MASTER_SKU_PATH)
    import datetime
    mtime = datetime.datetime.fromtimestamp(os.path.getmtime(MASTER_SKU_PATH), tz=IST).strftime('%d %b %Y, %H:%M IST')
    return jsonify({'ok': True, 'updated': mtime})


@app.route('/api/preflight', methods=['POST'])
def preflight():
    """
    Lightweight pre-sort scan. Called before the user confirms sorting.
    1. Detects accounts + label counts from uploaded PDFs (Feature 1).
    2. Checks for duplicate Order IDs against the 72-hour DB (Feature 2).
    Returns JSON the frontend uses to render confirmation modals.
    """
    req_tmp = tempfile.mkdtemp(prefix='fk_pre_')
    try:
        pdf_files = request.files.getlist('pdfs')
        if not pdf_files:
            return jsonify({'error': 'No PDF files provided'}), 400

        # Save uploads
        pdf_paths = []
        for upload in pdf_files:
            safe_fname = re.sub(r'[^A-Za-z0-9._-]', '_', os.path.basename(upload.filename))
            path = os.path.join(req_tmp, safe_fname)
            upload.save(path)
            pdf_paths.append(path)

        # Scan pages — extract account, order ID per page
        account_counts  = defaultdict(int)           # account -> label count
        account_oid_list = defaultdict(list)         # account -> [oid, ...] (allows dupes)

        for pdf_path in pdf_paths:
            reader = PdfReader(pdf_path)
            for page in reader.pages:
                text    = page.extract_text() or ''
                account = detect_account(text)
                account_counts[account] += 1
                for key in extract_label_keys(text):
                    # key = (order_id, awb) — same OID + different AWB = distinct label
                    account_oid_list[account].append(key)
            del reader
            gc.collect()

        # Detect within-batch duplicates: same (OID, AWB) pair seen more than once
        # Same OID but different AWB = different label = NOT a duplicate
        batch_dupes = {}
        for account, keys in account_oid_list.items():
            seen = set(); dupes = set()
            for key in keys:
                if key in seen:
                    dupes.add(key[0])   # report the OID (AWB implied same)
                seen.add(key)
            if dupes:
                batch_dupes[account] = sorted(dupes)

        # Unique (OID, AWB) keys per account — for DB check extract just OIDs
        # but only flag as DB-dupe if the exact (OID, AWB) combo was sorted before
        account_oids = {acc: set(k[0] for k in keys) for acc, keys in account_oid_list.items()}
        account_keys = {acc: set(keys) for acc, keys in account_oid_list.items()}

        # Check duplicates against 72h DB (stored as OID -> {awb: timestamp})
        db = _load_order_db()
        db_dupes = {}
        for account, keys in account_keys.items():
            existing = db.get(account, {})
            dupes = []
            for oid, awb in keys:
                if oid in existing:
                    # Check if this specific AWB was already sorted
                    stored = existing[oid]
                    if isinstance(stored, dict):
                        if awb in stored:   # exact (OID, AWB) match
                            dupes.append(oid)
                    else:
                        # Legacy format (timestamp only, no AWB stored) — flag by OID
                        dupes.append(oid)
            if dupes:
                db_dupes[account] = sorted(set(dupes))

        # Merge both duplicate sources — batch dupes take priority in the warning
        duplicates = {}
        all_dup_accounts = set(list(batch_dupes.keys()) + list(db_dupes.keys()))
        for account in all_dup_accounts:
            combined = sorted(set(batch_dupes.get(account, []) + db_dupes.get(account, [])))
            duplicates[account] = combined

        # Build account summary in KNOWN_ACCOUNTS order, unknowns appended
        ordered_accounts = []
        for acc in KNOWN_ACCOUNTS:
            if acc in account_counts:
                ordered_accounts.append({
                    'account': acc,
                    'label_count': account_counts[acc],
                    'order_count': len(account_oids[acc]),
                })
        for acc in account_counts:
            if acc not in KNOWN_ACCOUNTS:
                ordered_accounts.append({
                    'account': acc,
                    'label_count': account_counts[acc],
                    'order_count': len(account_oids[acc]),
                })

        return jsonify({
            'accounts':    ordered_accounts,
            'duplicates':  db_dupes,           # only historical — shown in popup
            'batch_dupes': batch_dupes,         # within-batch — auto-excluded silently
            'has_dupes':   bool(db_dupes),      # popup only for historical dupes
            'has_batch_dupes': bool(batch_dupes),
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        shutil.rmtree(req_tmp, ignore_errors=True)
        gc.collect()

@app.route('/api/sort', methods=['POST'])
def sort_labels():
    # Per-request temp dir — cleaned up on completion regardless of success/failure
    req_tmp = tempfile.mkdtemp(prefix='fk_sort_')
    try:
        pdf_files = request.files.getlist('pdfs')
        if not pdf_files:
            return jsonify({'error': 'No PDF files provided'}), 400
        if not os.path.exists(MASTER_SKU_PATH):
            return jsonify({'error': 'No Master SKU file found on server. Please upload one first.'}), 400

        # Excluded order IDs sent by frontend when user chose "Remove Duplicates"
        # Format: JSON-encoded dict {account: [oid, ...]}
        exclude_raw   = request.form.get('exclude_order_ids', '{}')
        try:
            exclude_map = json.loads(exclude_raw)  # {account: [oid, ...]} — user-chosen DB dupes
        except Exception:
            exclude_map = {}

        batch_exclude_raw = request.form.get('batch_exclude_oids', '{}')
        try:
            batch_exclude_map = json.loads(batch_exclude_raw)  # {account: [oid, ...]} — auto batch dupes
        except Exception:
            batch_exclude_map = {}

        # Only exclude historical dupes chosen by user — within-batch dedup is per-page below
        exclude_oids = set(oid for oids in exclude_map.values() for oid in oids)

        _seen_batch_keys = set()  # tracks (oid_tuple, awb) to keep first occurrence, skip repeats

        # ── Load master SKU (small xlsx, fine in RAM) ──────────────────────────
        with open(MASTER_SKU_PATH, 'rb') as f:
            master = load_sku_master(f.read())
        gc.collect()

        # ── Save uploaded PDFs to per-request temp dir ─────────────────────────
        pdf_paths = []
        for upload in pdf_files:
            # Sanitise filename
            safe_fname = re.sub(r'[^A-Za-z0-9._-]', '_', os.path.basename(upload.filename))
            path = os.path.join(req_tmp, safe_fname)
            upload.save(path)
            pdf_paths.append(path)

        # ── Step 1: Extract text using pypdf (no rendering = low RAM) ──────────
        # pypdf extracts embedded text directly — 10-20x less memory than pdfplumber
        account_pages = defaultdict(list)
        for pdf_path in pdf_paths:
            reader = PdfReader(pdf_path)
            for i, page in enumerate(reader.pages):
                text = page.extract_text() or ''
                account = detect_account(text)
                skus = extract_skus_from_page(text)
                order_ids  = extract_order_ids(text)
                page_awb   = extract_awb(text)
                # Skip page if its OID is in the historical exclude set (user chose to remove)
                if exclude_oids and any(oid in exclude_oids for oid in order_ids):
                    del text
                    continue
                # Within-batch dedup: keep first occurrence of each (oid, awb), skip repeats
                page_key = (tuple(sorted(order_ids)), page_awb)
                if page_key in _seen_batch_keys:
                    del text
                    continue
                _seen_batch_keys.add(page_key)
                account_pages[account].append({
                    'orig_path': pdf_path,
                    'orig_idx': i,
                    'skus': skus,
                    'text': text,
                    'order_ids': order_ids,
                })
                del text
            del reader   # release file handle + internal buffers immediately
            gc.collect()

        # ── Step 2: Per-account: classify → sort → write output PDFs ──────────
        output_files = []
        all_account_data = {}  # {account: {normal, dual, mixed, unknown}} for consolidated

        total_pages = sum(len(v) for v in account_pages.values())
        if not account_pages or total_pages == 0:
            return jsonify({'error': 'All labels were excluded as duplicates. This usually means the uploaded PDFs are identical to each other (same-batch duplicates). Nothing to sort.'}), 400

        for account, pages in account_pages.items():
            account_sku_map = get_account_master(master, account)
            normal, dual, mixed, unknown = classify_pages(pages, account_sku_map)
            sorted_normal = sort_normal(normal)
            sorted_dual   = sort_normal(dual)
            ordered = sorted_normal + sorted_dual + unknown + mixed

            # Store for consolidated PDF
            all_account_data[account] = {
                'normal':  sorted_normal + sorted_dual,
                'dual':    [],   # already merged into normal above
                'mixed':   mixed,
                'unknown': unknown,
            }

            safe_name = re.sub(r'[^A-Za-z0-9_]', '_', account)
            labels_path = os.path.join(req_tmp, f'{safe_name}_sorted_labels.pdf')

            # Build sorted labels PDF one source-PDF at a time to cap peak RAM.
            # We open each source PDF, copy its needed pages, then close it before
            # moving to the next — so at most ONE source PDF is open at a time.
            pages_by_pdf = defaultdict(list)  # {src_path: [(output_pos, page_idx)]}
            for pos, pd_item in enumerate(ordered):
                pages_by_pdf[pd_item['orig_path']].append((pos, pd_item['orig_idx']))

            page_slots = [None] * len(ordered)
            for pdf_src, idx_pairs in pages_by_pdf.items():
                reader = PdfReader(pdf_src)
                for pos, page_idx in idx_pairs:
                    page_slots[pos] = reader.pages[page_idx]
                # NOTE: We intentionally keep reader alive until after writer.write()
                # because pypdf pages hold a reference back to their reader.
                # We collect all readers, write the PDF, then delete them all.
                pages_by_pdf[pdf_src] = (reader, idx_pairs)  # store reader ref

            writer = PdfWriter()
            for page in page_slots:
                writer.add_page(page)
            with open(labels_path, 'wb') as f:
                writer.write(f)

            # Now safe to free everything
            del writer, page_slots
            for pdf_src, (reader, _) in pages_by_pdf.items():
                del reader
            del pages_by_pdf
            gc.collect()

            # Build summary PDF
            summary_path = os.path.join(req_tmp, f'{safe_name}_summary.pdf')
            build_summary_pdf(account, normal, dual, mixed, unknown, account_sku_map, summary_path)

            # Persist to Railway Volume (atomic: copy then rename avoids partial writes)
            persist_labels  = os.path.join(OUTPUT_DIR, f'{safe_name}_labels.pdf')
            persist_summary = os.path.join(OUTPUT_DIR, f'{safe_name}_summary.pdf')
            _atomic_copy(labels_path,  persist_labels)
            _atomic_copy(summary_path, persist_summary)

            # Telegram notification — send both PDFs immediately after persisting
            try:
                tg_notify_sort_done(
                    account=account,
                    total=len(pages),
                    normal_count=len(normal),
                    mixed_count=len(mixed),
                    unknown_count=len(unknown),
                    labels_path=persist_labels,
                    summary_path=persist_summary,
                )
            except Exception as tg_err:
                print(f'[Telegram notify] {tg_err}')

            # Update metadata JSON
            meta = {}
            if os.path.exists(OUTPUTS_META):
                with open(OUTPUTS_META, 'r') as mf:
                    try:
                        meta = json.load(mf)
                    except json.JSONDecodeError:
                        meta = {}
            meta[account] = {
                'timestamp': datetime.datetime.now(tz=IST).strftime('%d %b %Y, %H:%M IST'),
                'total': len(pages),
                'sku_count': len(normal),
                'labels_file': f'{safe_name}_labels.pdf',
                'summary_file': f'{safe_name}_summary.pdf',
            }
            with open(OUTPUTS_META, 'w') as mf:
                json.dump(meta, mf)

            output_files.append({
                'name': account,
                'total': len(pages),
                'sku_count': len(normal),
                'labels_file': f'{safe_name}_labels.pdf',
                'summary_file': f'{safe_name}_summary.pdf',
            })

            # Free page data before next account
            del normal, dual, mixed, unknown, ordered, pages
            gc.collect()

        # ── Step 3: Build consolidated PDF if multiple accounts detected ──────
        active_accounts = [a for a in ACCOUNT_ORDER if a in all_account_data]
        if len(active_accounts) >= 1:
            consolidated_labels_path  = os.path.join(req_tmp, 'consolidated_labels.pdf')
            consolidated_summary_path = os.path.join(req_tmp, 'consolidated_summary.pdf')
            build_consolidated_pdf(all_account_data, consolidated_labels_path)
            build_consolidated_summary_pdf(all_account_data, consolidated_summary_path)

            persist_con_labels  = os.path.join(OUTPUT_DIR, 'consolidated_labels.pdf')
            persist_con_summary = os.path.join(OUTPUT_DIR, 'consolidated_summary.pdf')
            _atomic_copy(consolidated_labels_path,  persist_con_labels)
            _atomic_copy(consolidated_summary_path, persist_con_summary)

            # Update meta with consolidated entry
            meta = {}
            if os.path.exists(OUTPUTS_META):
                with open(OUTPUTS_META, 'r') as mf:
                    try: meta = json.load(mf)
                    except: pass
            meta['__consolidated__'] = {
                'timestamp': datetime.datetime.now(tz=IST).strftime('%d %b %Y, %H:%M IST'),
                'accounts':  active_accounts,
                'labels_file':  'consolidated_labels.pdf',
                'summary_file': 'consolidated_summary.pdf',
            }
            with open(OUTPUTS_META, 'w') as mf:
                json.dump(meta, mf)

        # ── Record sorted Order IDs into 72h DB ──────────────────────────────
        account_sorted_keys = {}
        for account, pages in account_pages.items():
            keys = set()
            for pg in pages:
                awb = extract_awb(pg.get('text', '') or '')
                for oid in pg.get('order_ids', []):
                    keys.add((oid, awb))
            if keys:
                account_sorted_keys[account] = list(keys)
        if account_sorted_keys:
            try:
                _record_order_ids(account_sorted_keys)
            except Exception as db_err:
                print(f'[OrderDB] Failed to record: {db_err}')

        return jsonify({'accounts': output_files})

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

    finally:
        # Always clean up the per-request temp dir — critical for RAM & disk
        try:
            shutil.rmtree(req_tmp, ignore_errors=True)
        except Exception:
            pass
        gc.collect()


@app.route('/api/debug-pdf', methods=['POST'])
def debug_pdf():
    """Upload one PDF and return raw extracted text per page — for diagnosing SKU extraction."""
    pdf_file = request.files.get('pdf')
    if not pdf_file:
        return jsonify({'error': 'No PDF provided'}), 400
    req_tmp = tempfile.mkdtemp(prefix='fk_dbg_')
    try:
        path = os.path.join(req_tmp, 'debug.pdf')
        pdf_file.save(path)
        reader = PdfReader(path)
        pages_out = []
        for i, page in enumerate(reader.pages[:5]):  # first 5 pages only
            text = page.extract_text() or ''
            lines = text.split('\n')
            pages_out.append({
                'page': i,
                'raw_text': text,
                'lines': lines,
                'account_detected': detect_account(text),
                'skus_found': extract_skus_from_page(text),
            })
        return jsonify({'pages': pages_out})
    finally:
        shutil.rmtree(req_tmp, ignore_errors=True)


# ─────────────────────────────────────────────
# LISTINGS API — FETCH / UPDATE PRICE / UPDATE INVENTORY
# ─────────────────────────────────────────────

@app.route('/api/listings/<account>', methods=['GET'])
def listings_get(account):
    """
    Fetch one page of ACTIVE listings for the given account from Flipkart API.
    Query params:
      - page_id: encrypted page cursor from previous response (omit for first page)
    Step 1: POST /sellers/listings/v3/search  — one page (500 SKUs), returns sku_id + product_id
    Step 2: POST /sellers/listings/v3/details — batches of 10, returns price + inventory
    Returns: { listings, next_page_id, has_more, total_this_page }
    """
    import requests as _req
    account = account.upper().replace('-', ' ')
    if account not in KNOWN_ACCOUNTS:
        return jsonify({'error': f'Unknown account: {account}'}), 400
    page_id = request.args.get('page_id', None) or None   # None means first page
    try:
        token   = fk_get_token(account)
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type':  'application/json',
        }

        # ── Step 1: fetch ONE search page (up to 500 sku_ids) ──
        search_url  = f'{FK_API_BASE}/sellers/listings/v3/search'
        payload     = {'filters': {'listing_status': 'ACTIVE'}, 'page_id': page_id}
        r = _req.post(search_url, json=payload, headers=headers, timeout=30)
        if r.status_code != 200:
            return jsonify({'error': f'Flipkart API error [{r.status_code}]: {r.text[:300]}'}), 502
        data = r.json()

        raw_listings = data.get('listings', [])
        sku_entries  = []
        if isinstance(raw_listings, list):
            for item in raw_listings:
                sid = item.get('sku_id') or item.get('skuId', '')
                if sid:
                    sku_entries.append({'sku_id': sid, 'product_id': item.get('product_id') or item.get('productId', '')})
        elif isinstance(raw_listings, dict):
            for sid, item in raw_listings.items():
                sku_entries.append({'sku_id': sid, 'product_id': item.get('product_id') or item.get('productId', '')})

        next_page_id = data.get('next_page_id') or None
        has_more     = bool(data.get('has_more')) and bool(next_page_id)

        if not sku_entries:
            return jsonify({'ok': True, 'account': account, 'listings': [],
                            'next_page_id': None, 'has_more': False, 'total_this_page': 0})

        # ── Step 2: fetch details in batches of 10 ──
        details_url = f'{FK_API_BASE}/sellers/listings/v3/details'
        detail_map  = {}
        sku_ids     = [e['sku_id'] for e in sku_entries]
        for i in range(0, len(sku_ids), 10):
            batch = sku_ids[i:i+10]
            dr = _req.post(details_url, json={'sku_ids': batch}, headers=headers, timeout=30)
            if dr.status_code == 200:
                for sku_id, det in dr.json().get('available', {}).items():
                    price = det.get('price', {})
                    locs  = det.get('locations', [])
                    stock = sum(loc.get('inventory', 0) for loc in locs if loc.get('status') == 'ENABLED')
                    detail_map[sku_id] = {
                        'selling_price': price.get('selling_price', ''),
                        'mrp':           price.get('mrp', ''),
                        'mop':           price.get('mop', ''),
                        'stock_count':   stock,
                        'fsn':           det.get('fsn', ''),
                        'locations':     [{'id': loc.get('id',''), 'status': loc.get('status','ENABLED')} for loc in locs],
                    }

        # ── Merge ──
        listings = []
        for e in sku_entries:
            sid = e['sku_id']
            det = detail_map.get(sid, {})
            listings.append({
                'sku_id':        sid,
                'product_id':    e['product_id'],
                'fsn':           det.get('fsn', ''),
                'selling_price': det.get('selling_price', ''),
                'mrp':           det.get('mrp', ''),
                'mop':           det.get('mop', ''),
                'stock_count':   det.get('stock_count', 0),
                'locations':     det.get('locations', []),
            })

        return jsonify({
            'ok':             True,
            'account':        account,
            'listings':       listings,
            'next_page_id':   next_page_id,
            'has_more':       has_more,
            'total_this_page': len(listings),
        })
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/listings/<account>/update-price', methods=['POST'])
def listings_update_price(account):
    """
    Update selling price for one or more SKUs.
    Body: { skus: [{sku_id, product_id, mrp, selling_price}, ...] }
    Batches into groups of 10 (API limit). Returns per-SKU Flipkart status.
    """
    import requests as _req
    account = account.upper().replace('-', ' ')
    if account not in KNOWN_ACCOUNTS:
        return jsonify({'error': f'Unknown account: {account}'}), 400
    body = request.get_json() or {}
    skus = body.get('skus', [])
    if not skus:
        return jsonify({'error': 'No SKUs provided'}), 400
    try:
        token   = fk_get_token(account)
        headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
        url     = f'{FK_API_BASE}/sellers/listings/v3/update/price'
        results = {}
        for i in range(0, len(skus), 10):
            batch   = skus[i:i+10]
            payload = {}
            for s in batch:
                payload[s['sku_id']] = {
                    'product_id': s['product_id'],
                    'price': {
                        'mrp':      int(s['mrp']),
                        'mop':      int(s['mop']),
                        'currency': 'INR',
                    }
                }
            r = _req.post(url, json=payload, headers=headers, timeout=30)
            if r.status_code == 200:
                results.update(r.json())
            else:
                for s in batch:
                    results[s['sku_id']] = {
                        'status': 'failure',
                        'errors': [{'severity': 'ERROR', 'description': f'HTTP {r.status_code}: {r.text[:200]}'}]
                    }
        return jsonify({'ok': True, 'results': results})
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/listings/<account>/update-inventory', methods=['POST'])
def listings_update_inventory(account):
    """
    Update stock count for one or more SKUs.
    Body: { skus: [{sku_id, stock_count}, ...] }
    Batches into groups of 10. Returns per-SKU Flipkart status.
    """
    import requests as _req
    account = account.upper().replace('-', ' ')
    if account not in KNOWN_ACCOUNTS:
        return jsonify({'error': f'Unknown account: {account}'}), 400
    body = request.get_json() or {}
    skus = body.get('skus', [])
    if not skus:
        return jsonify({'error': 'No SKUs provided'}), 400
    try:
        token   = fk_get_token(account)
        headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
        url     = f'{FK_API_BASE}/sellers/listings/v3/update/inventory'
        results = {}
        for i in range(0, len(skus), 10):
            batch   = skus[i:i+10]
            payload = {}
            for s in batch:
                # Use stored location IDs if available, otherwise fall back to a single
                # default location entry. The API requires locations[].id + status + inventory.
                loc_list = s.get('locations') or []
                if loc_list:
                    locations = [{'id': loc['id'], 'inventory': int(s['stock_count'])} for loc in loc_list]
                else:
                    locations = [{'id': 'default', 'inventory': int(s['stock_count'])}]
                payload[s['sku_id']] = {
                    'product_id': s.get('product_id', ''),
                    'locations':  locations,
                }
            r = _req.post(url, json=payload, headers=headers, timeout=30)
            if r.status_code == 200:
                results.update(r.json())
            else:
                for s in batch:
                    results[s['sku_id']] = {
                        'status': 'failure',
                        'errors': [{'severity': 'ERROR', 'description': f'HTTP {r.status_code}: {r.text[:200]}'}]
                    }
        return jsonify({'ok': True, 'results': results})
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/listings/<account>/by-skus', methods=['POST'])
def listings_by_skus(account):
    """
    Fetch listing details for a specific list of SKU IDs.
    Body: { sku_ids: [...] }
    Calls /listings/v3/details in batches of 10.
    Used by the product filter to load only the relevant SKUs.
    """
    import requests as _req
    account = account.upper().replace('-', ' ')
    if account not in KNOWN_ACCOUNTS:
        return jsonify({'error': f'Unknown account: {account}'}), 400
    body    = request.get_json() or {}
    sku_ids = body.get('sku_ids', [])
    if not sku_ids:
        return jsonify({'ok': True, 'listings': [], 'total': 0})
    try:
        token   = fk_get_token(account)
        headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
        details_url = f'{FK_API_BASE}/sellers/listings/v3/details'
        detail_map  = {}
        for i in range(0, len(sku_ids), 10):
            batch = sku_ids[i:i+10]
            dr = _req.post(details_url, json={'sku_ids': batch}, headers=headers, timeout=30)
            if dr.status_code == 200:
                for sku_id, det in dr.json().get('available', {}).items():
                    price = det.get('price', {})
                    locs  = det.get('locations', [])
                    stock = sum(loc.get('inventory', 0) for loc in locs if loc.get('status') == 'ENABLED')
                    detail_map[sku_id] = {
                        'selling_price': price.get('selling_price', ''),
                        'mrp':           price.get('mrp', ''),
                        'mop':           price.get('mop', ''),
                        'stock_count':   stock,
                        'fsn':           det.get('fsn', ''),
                        'locations':     [{'id': loc.get('id', '')} for loc in locs],
                        'product_id':    det.get('product_id', ''),
                    }
        listings = []
        for sku_id in sku_ids:
            det = detail_map.get(sku_id, {})
            listings.append({
                'sku_id':        sku_id,
                'product_id':    det.get('product_id', ''),
                'fsn':           det.get('fsn', ''),
                'selling_price': det.get('selling_price', ''),
                'mrp':           det.get('mrp', ''),
                'mop':           det.get('mop', ''),
                'stock_count':   det.get('stock_count', 0),
                'locations':     det.get('locations', []),
            })
        return jsonify({'ok': True, 'listings': listings, 'total': len(listings)})
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ─────────────────────────────────────────────
# AUTO-DISPATCH PIPELINE
# Runs on cron: 8:30am, 11:30am, 2pm, 5pm IST
# For Cutest Club only (the only API-configured account).
#
# Flow:
#   1. Fetch all APPROVED shipments
#   2. Trigger label generation (pack) with default/existing dimensions
#   3. Download combined labels PDF from API
#   4. Run through Label Sorter pipeline → save to OUTPUT_DIR
#   5. Mark all as READY_TO_DISPATCH
#   6. Telegram notification + update OUTPUTS_META
# ─────────────────────────────────────────────

FK_AUTO_DISPATCH_ACCOUNT  = 'CUTEST CLUB'
FK_AUTO_DISPATCH_LOCATION = 'LOC87f71f39207645b9b9427c976d4a7da1'

# Default package dimensions for orders with no pre-existing dimensions
FK_DEFAULT_DIMS = {'length': 10, 'breadth': 5, 'height': 5, 'weight': 0.2}


def _fk_auto_dispatch():
    """
    Core auto-dispatch pipeline. Called by the cron route and the
    manual trigger route. Returns a result dict.
    """
    import requests as _req

    account  = FK_AUTO_DISPATCH_ACCOUNT
    location = FK_AUTO_DISPATCH_LOCATION

    try:
        token   = fk_get_token(account)
    except RuntimeError as e:
        return {'ok': False, 'error': str(e)}

    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    base    = FK_API_BASE + '/sellers'

    # ── Step 1: Fetch all APPROVED shipments ──────────────────────────────────
    approved_shipments = []   # list of {shipmentId, subShipments}
    next_url = f'{base}/v3/shipments/filter/'
    payload  = {
        'filter': {
            'type':   'preDispatch',
            'states': ['APPROVED'],
            'locationId': location,
        },
        'pagination': {'pageSize': 20},
    }
    fetched = 0
    while next_url and fetched < 2000:
        try:
            r = (_req.post(next_url, json=payload, headers=headers, timeout=30)
                 if next_url.endswith('/filter/')
                 else _req.get(next_url, headers=headers, timeout=30))
            if r.status_code != 200:
                return {'ok': False, 'error': f'Shipment filter failed [{r.status_code}]: {r.text[:200]}'}
            data = r.json()
            for s in data.get('shipments', []):
                approved_shipments.append(s)
                fetched += 1
            if not data.get('hasMore') or not data.get('nextPageUrl'):
                break
            raw_next = data['nextPageUrl']
            # Flipkart sometimes returns a relative path — prepend base if needed
            if raw_next.startswith('/'):
                next_url = f'{FK_API_BASE}/sellers{raw_next}' if not raw_next.startswith('/sellers') else f'{FK_API_BASE}{raw_next}'
            elif not raw_next.startswith('http'):
                next_url = f'{FK_API_BASE}/sellers/{raw_next}'
            else:
                next_url = raw_next
            payload  = None   # subsequent pages use GET to nextPageUrl
        except Exception as e:
            return {'ok': False, 'error': f'Shipment filter error: {e}'}

    if not approved_shipments:
        return {'ok': True, 'message': 'No APPROVED shipments found — nothing to process.',
                'total': 0, 'packed': 0, 'rtd': 0}

    print(f'[AutoDispatch] Found {len(approved_shipments)} APPROVED shipments for {account}')

    # ── Step 2: Pack (trigger label generation) in batches of 25 ─────────────
    pack_url   = f'{base}/v3/shipments/labels'
    shipment_ids_packed = []
    pack_errors = []

    for i in range(0, len(approved_shipments), 25):
        batch = approved_shipments[i:i+25]
        pack_payload = {'shipments': []}
        for s in batch:
            sid = s.get('shipmentId') or s.get('shipment_id', '')
            # Build sub-shipment entries with default dimensions
            sub_shipments = []
            for sub in s.get('subShipments', [s]):   # fall back to shipment itself
                sub_id = sub.get('subShipmentId') or sub.get('shipmentId') or sid
                sub_shipments.append({
                    'subShipmentId': sub_id,
                    'dimensions': FK_DEFAULT_DIMS,
                })
            pack_payload['shipments'].append({
                'shipmentId':  sid,
                'locationId':  location,
                'subShipments': sub_shipments,
            })
        try:
            r = _req.post(pack_url, json=pack_payload, headers=headers, timeout=60)
            if r.status_code == 200:
                for result in r.json().get('shipments', []):
                    if result.get('status', '').upper() in ('SUCCESS', 'PACKED', ''):
                        shipment_ids_packed.append(result['shipmentId'])
                    else:
                        pack_errors.append(f"{result['shipmentId']}: {result.get('errorMessage','')}")
            else:
                pack_errors.append(f'Batch {i//25+1} HTTP {r.status_code}: {r.text[:150]}')
        except Exception as e:
            pack_errors.append(f'Batch {i//25+1} error: {e}')

    # Include any IDs from original list not explicitly failed (API may return empty list on full success)
    if not shipment_ids_packed:
        shipment_ids_packed = [s.get('shipmentId') or s.get('shipment_id', '')
                               for s in approved_shipments]

    print(f'[AutoDispatch] Packed {len(shipment_ids_packed)} shipments. Errors: {len(pack_errors)}')
    if pack_errors:
        for e in pack_errors[:5]:
            print(f'  [PackError] {e}')

    if not shipment_ids_packed:
        return {'ok': False, 'error': 'All shipments failed to pack.',
                'pack_errors': pack_errors}

    # ── Step 3: Download labels PDF ───────────────────────────────────────────
    # Flipkart accepts comma-separated shipment IDs in the URL path.
    # Batch into groups of 25 and concatenate the PDFs.
    label_pdf_parts = []
    for i in range(0, len(shipment_ids_packed), 25):
        batch_ids = ','.join(shipment_ids_packed[i:i+25])
        try:
            r = _req.get(
                f'{base}/v3/shipments/{batch_ids}/labels',
                headers={**headers, 'Accept': 'application/pdf'},
                timeout=60,
            )
            if r.status_code == 200 and r.content:
                label_pdf_parts.append(r.content)
            else:
                print(f'[AutoDispatch] Label download batch {i//25+1} failed [{r.status_code}]')
        except Exception as e:
            print(f'[AutoDispatch] Label download error: {e}')

    if not label_pdf_parts:
        return {'ok': False, 'error': 'Failed to download labels PDF from Flipkart API.',
                'pack_errors': pack_errors}

    # Merge all label PDF bytes into one (simple concatenation via PdfWriter)
    combined_writer = PdfWriter()
    for pdf_bytes in label_pdf_parts:
        try:
            reader = PdfReader(io.BytesIO(pdf_bytes))
            for page in reader.pages:
                combined_writer.add_page(page)
            del reader
        except Exception as e:
            print(f'[AutoDispatch] PDF merge error: {e}')

    # ── Step 4: Run Label Sorter pipeline ─────────────────────────────────────
    req_tmp = tempfile.mkdtemp(prefix='fk_auto_')
    try:
        # Save combined PDF to temp file
        tmp_pdf_path = os.path.join(req_tmp, 'auto_labels.pdf')
        with open(tmp_pdf_path, 'wb') as f:
            combined_writer.write(f)
        del combined_writer
        gc.collect()

        if not os.path.exists(MASTER_SKU_PATH):
            return {'ok': False, 'error': 'Master SKU file not found — cannot sort labels.'}

        with open(MASTER_SKU_PATH, 'rb') as f:
            master = load_sku_master(f.read())

        # Extract pages
        account_pages = defaultdict(list)
        seen_keys     = set()
        reader = PdfReader(tmp_pdf_path)
        for i, page in enumerate(reader.pages):
            text       = page.extract_text() or ''
            acct       = detect_account(text)
            skus       = extract_skus_from_page(text)
            order_ids  = extract_order_ids(text)
            page_awb   = extract_awb(text)
            page_key   = (tuple(sorted(order_ids)), page_awb)
            if page_key in seen_keys:
                del text; continue
            seen_keys.add(page_key)
            account_pages[acct].append({
                'orig_path': tmp_pdf_path,
                'orig_idx':  i,
                'skus':      skus,
                'text':      text,
                'order_ids': order_ids,
            })
            del text
        del reader
        gc.collect()

        sort_results = []
        for acct, pages in account_pages.items():
            acct_master = get_account_master(master, acct)
            normal, dual, mixed, unknown = classify_pages(pages, acct_master)
            sorted_normal = sort_normal(normal)
            sorted_dual   = sort_normal(dual)
            ordered       = sorted_normal + sorted_dual + unknown + mixed

            safe_name    = re.sub(r'[^A-Za-z0-9_]', '_', acct)
            labels_path  = os.path.join(req_tmp, f'{safe_name}_labels.pdf')
            summary_path = os.path.join(req_tmp, f'{safe_name}_summary.pdf')

            # Write sorted labels PDF
            pages_by_pdf = defaultdict(list)
            for pos, pd_item in enumerate(ordered):
                pages_by_pdf[pd_item['orig_path']].append((pos, pd_item['orig_idx']))
            page_slots = [None] * len(ordered)
            readers_open = {}
            for pdf_src, idx_pairs in pages_by_pdf.items():
                r2 = PdfReader(pdf_src)
                readers_open[pdf_src] = r2
                for pos, page_idx in idx_pairs:
                    page_slots[pos] = r2.pages[page_idx]
            writer = PdfWriter()
            for page in page_slots:
                writer.add_page(page)
            with open(labels_path, 'wb') as f:
                writer.write(f)
            del writer, page_slots
            for r2 in readers_open.values():
                del r2
            gc.collect()

            # Build summary
            build_summary_pdf(acct, normal, dual, mixed, unknown, acct_master, summary_path)

            # Persist to Railway Volume
            persist_labels  = os.path.join(OUTPUT_DIR, f'{safe_name}_labels.pdf')
            persist_summary = os.path.join(OUTPUT_DIR, f'{safe_name}_summary.pdf')
            _atomic_copy(labels_path,  persist_labels)
            _atomic_copy(summary_path, persist_summary)

            # Update OUTPUTS_META
            meta = {}
            if os.path.exists(OUTPUTS_META):
                with open(OUTPUTS_META, 'r') as mf:
                    try: meta = json.load(mf)
                    except: pass
            meta[acct] = {
                'timestamp':    datetime.datetime.now(tz=IST).strftime('%d %b %Y, %H:%M IST'),
                'total':        len(pages),
                'sku_count':    len(normal),
                'labels_file':  f'{safe_name}_labels.pdf',
                'summary_file': f'{safe_name}_summary.pdf',
            }
            with open(OUTPUTS_META, 'w') as mf:
                json.dump(meta, mf)

            # Telegram
            try:
                tg_notify_sort_done(
                    account=acct, total=len(pages),
                    normal_count=len(normal), mixed_count=len(mixed),
                    unknown_count=len(unknown),
                    labels_path=persist_labels, summary_path=persist_summary,
                )
            except Exception as tg_err:
                print(f'[AutoDispatch Telegram] {tg_err}')

            sort_results.append({
                'account':     acct,
                'total':       len(pages),
                'sku_count':   len(normal),
                'mixed':       len(mixed),
                'unknown':     len(unknown),
            })
            del normal, dual, mixed, unknown, ordered, pages
            gc.collect()

    finally:
        shutil.rmtree(req_tmp, ignore_errors=True)

    # ── Step 5: Mark RTD in batches of 25 ────────────────────────────────────
    dispatch_url = f'{base}/v3/shipments/dispatch'
    rtd_count    = 0
    rtd_errors   = []
    for i in range(0, len(shipment_ids_packed), 25):
        batch = shipment_ids_packed[i:i+25]
        try:
            r = _req.post(dispatch_url,
                          json={'shipmentIds': batch, 'locationId': location},
                          headers=headers, timeout=30)
            if r.status_code == 200:
                for result in r.json().get('shipments', []):
                    if result.get('status', '').upper() in ('SUCCESS', 'READY_TO_DISPATCH', ''):
                        rtd_count += 1
                    else:
                        rtd_errors.append(f"{result['shipmentId']}: {result.get('errorMessage','')}")
            else:
                rtd_errors.append(f'RTD batch {i//25+1} HTTP {r.status_code}: {r.text[:150]}')
        except Exception as e:
            rtd_errors.append(f'RTD batch {i//25+1} error: {e}')

    # If API returns empty shipments list, count all as RTD (some API versions do this on full success)
    if rtd_count == 0 and not rtd_errors:
        rtd_count = len(shipment_ids_packed)

    print(f'[AutoDispatch] RTD: {rtd_count} marked, {len(rtd_errors)} errors')

    timestamp = datetime.datetime.now(tz=IST).strftime('%d %b %Y, %H:%M IST')
    return {
        'ok':           True,
        'timestamp':    timestamp,
        'total':        len(approved_shipments),
        'packed':       len(shipment_ids_packed),
        'rtd':          rtd_count,
        'sort_results': sort_results,
        'pack_errors':  pack_errors[:10],
        'rtd_errors':   rtd_errors[:10],
    }



@app.route('/api/auto-dispatch/preview', methods=['POST'])
def auto_dispatch_preview():
    """
    Returns count of APPROVED shipments without doing anything.
    Used by the frontend confirmation step.
    """
    import requests as _req
    data = request.get_json() or {}
    if data.get('pin') != '848424':
        return jsonify({'error': 'Invalid PIN'}), 403

    account  = FK_AUTO_DISPATCH_ACCOUNT
    location = FK_AUTO_DISPATCH_LOCATION
    try:
        token = fk_get_token(account)
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500

    headers  = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    base     = FK_API_BASE + '/sellers'
    count    = 0
    next_url = f'{base}/v3/shipments/filter/'
    payload  = {
        'filter': {
            'type':      'preDispatch',
            'states':    ['APPROVED'],
            'locationId': location,
        },
        'pagination': {'pageSize': 20},
    }
    try:
        while next_url and count < 2000:
            r = (_req.post(next_url, json=payload, headers=headers, timeout=30)
                 if next_url.endswith('/filter/')
                 else _req.get(next_url, headers=headers, timeout=30))
            if r.status_code != 200:
                return jsonify({'error': f'API error [{r.status_code}]: {r.text[:200]}'}), 502
            resp_data = r.json()
            count += len(resp_data.get('shipments', []))
            if not resp_data.get('hasMore') or not resp_data.get('nextPageUrl'):
                break
            raw_next = resp_data['nextPageUrl']
            if raw_next.startswith('/'):
                next_url = f'{FK_API_BASE}/sellers{raw_next}' if not raw_next.startswith('/sellers') else f'{FK_API_BASE}{raw_next}'
            elif not raw_next.startswith('http'):
                next_url = f'{FK_API_BASE}/sellers/{raw_next}'
            else:
                next_url = raw_next
            payload = None
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    return jsonify({'ok': True, 'count': count})

@app.route('/api/auto-dispatch', methods=['POST'])
def auto_dispatch_trigger():
    """
    Manual trigger for the auto-dispatch pipeline.
    Also called by Railway cron at 08:30, 11:30, 14:00, 17:00 IST.
    PIN-protected to prevent accidental triggers.
    """
    data = request.get_json() or {}
    if data.get('pin') != '848424':
        return jsonify({'error': 'Invalid PIN'}), 403
    result = _fk_auto_dispatch()
    return jsonify(result)


@app.route('/api/auto-dispatch/cron', methods=['GET', 'POST'])
def auto_dispatch_cron():
    """
    Called by Railway cron job — no PIN required (Railway internal only).
    Logs result to stdout (visible in Railway logs).
    """
    print(f'[AutoDispatch Cron] Triggered at {datetime.datetime.now(tz=IST).strftime("%d %b %Y, %H:%M IST")}')
    result = _fk_auto_dispatch()
    print(f'[AutoDispatch Cron] Result: {json.dumps(result)}')
    return jsonify(result)

@app.route('/api/debug')
def debug():
    import datetime
    info = {
        'data_dir': _data_dir,
        'data_dir_exists': os.path.isdir(_data_dir),
        'output_dir': OUTPUT_DIR,
        'output_dir_exists': os.path.isdir(OUTPUT_DIR),
        'outputs_meta_exists': os.path.exists(OUTPUTS_META),
        'master_exists': os.path.exists(MASTER_SKU_PATH),
        'output_dir_files': os.listdir(OUTPUT_DIR) if os.path.isdir(OUTPUT_DIR) else [],
        'data_dir_files': os.listdir(_data_dir) if os.path.isdir(_data_dir) else [],
    }
    if os.path.exists(OUTPUTS_META):
        with open(OUTPUTS_META, 'r') as f:
            info['meta_contents'] = json.load(f)
    return jsonify(info)


@app.route('/api/fk-token-status')
def fk_token_status():
    """
    Check Flipkart API token status for all configured accounts.
    Safe to expose — returns no secrets, just connectivity status.
    """
    results = {}
    for account in KNOWN_ACCOUNTS:
        app_id, app_secret = _fk_credentials(account)
        if not app_id:
            results[account] = {'configured': False}
            continue
        cache = _fk_load_token_cache()
        entry = cache.get(account, {})
        now   = datetime.datetime.now(tz=datetime.timezone.utc).timestamp()
        expires_at = entry.get('expires_at', 0)
        hours_left = max(0, int((expires_at - now) / 3600)) if expires_at else 0
        results[account] = {
            'configured':  True,
            'has_token':   bool(entry.get('access_token')),
            'fetched_at':  entry.get('fetched_at', 'never'),
            'hours_left':  hours_left,
        }
    return jsonify(results)


@app.route('/api/fk-token-refresh', methods=['POST'])
def fk_token_refresh():
    """Force-refresh the Flipkart token for a given account. PIN protected."""
    data    = request.get_json() or {}
    if data.get('pin') != '848424':
        return jsonify({'error': 'Invalid PIN'}), 403
    account = data.get('account', '').upper()
    if account not in KNOWN_ACCOUNTS:
        return jsonify({'error': f'Unknown account: {account}'}), 400
    # Clear cached token to force refresh
    cache = _fk_load_token_cache()
    cache.pop(account, None)
    _fk_save_token_cache(cache)
    try:
        token = fk_get_token(account)
        return jsonify({'ok': True, 'account': account, 'token_preview': token[:8] + '…'})
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/latest-outputs')
def latest_outputs():
    meta = {}
    if os.path.exists(OUTPUTS_META):
        with open(OUTPUTS_META, 'r') as f:
            meta = json.load(f)
    result = []
    for account in KNOWN_ACCOUNTS:
        if account in meta:
            entry = meta[account].copy()
            entry['account'] = account
            entry['has_output'] = True
        else:
            entry = {'account': account, 'has_output': False}
        result.append(entry)
    # Consolidated entry
    con = meta.get('__consolidated__')
    consolidated = {
        'has_output': bool(con),
        'timestamp':  con['timestamp']  if con else None,
        'accounts':   con['accounts']   if con else [],
        'labels_file':  con['labels_file']  if con else None,
        'summary_file': con['summary_file'] if con else None,
    }
    return jsonify({'outputs': result, 'consolidated': consolidated})


@app.route('/api/download-master')
def download_master():
    if not os.path.exists(MASTER_SKU_PATH):
        return "Master SKU file not found", 404
    return send_file(MASTER_SKU_PATH, as_attachment=True, download_name='master_sku.xlsx')


@app.route('/api/sku-data')
def sku_data():
    """Return all SKU data from master xlsx as JSON, keyed by sheet/account."""
    if not os.path.exists(MASTER_SKU_PATH):
        return jsonify({'error': 'No master SKU file found'}), 404
    wb = openpyxl.load_workbook(MASTER_SKU_PATH)
    result = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # skip header
            sku, p1, s1, p2, s2 = (row[0], row[1], row[2], row[3], row[4]) if len(row) >= 5 else (row[0], row[1], row[2], None, None)
            if not sku and not p1:
                continue
            rows.append({
                'sku': str(sku).strip() if sku else '',
                'p1': str(p1).strip() if p1 else '',
                's1': int(s1) if s1 else 1,
                'p2': str(p2).strip() if p2 else '',
                's2': int(s2) if s2 else '',
            })
        result[sheet] = rows
    # Collect all unique product names across all sheets
    all_products = set()
    for rows in result.values():
        for r in rows:
            if r['p1']: all_products.add(r['p1'])
            if r['p2']: all_products.add(r['p2'])
    return jsonify({'sheets': result, 'products': sorted(all_products)})


@app.route('/api/sku-save', methods=['POST'])
def sku_save():
    """Save edited SKU data back to master xlsx. Requires PIN."""
    data = request.get_json()
    if not data or data.get('pin') != '848424':
        return jsonify({'error': 'Invalid PIN'}), 403
    sheets_data = data.get('sheets', {})
    if not os.path.exists(MASTER_SKU_PATH):
        return jsonify({'error': 'No master SKU file found'}), 404
    wb = openpyxl.load_workbook(MASTER_SKU_PATH)
    for sheet_name, rows in sheets_data.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Clear existing data rows (keep header row 1)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
        # Write new rows
        for i, row in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=row.get('sku', ''))
            ws.cell(row=i, column=2, value=row.get('p1', ''))
            ws.cell(row=i, column=3, value=int(row['s1']) if row.get('s1') else 1)
            ws.cell(row=i, column=4, value=row.get('p2', '') or None)
            ws.cell(row=i, column=5, value=int(row['s2']) if row.get('s2') else None)
    wb.save(MASTER_SKU_PATH)
    return jsonify({'ok': True})


@app.route('/api/download/<filename>')
def download(filename):
    # Only serve from the persistent output dir (temp files are cleaned up after each request)
    persist_path = os.path.join(OUTPUT_DIR, filename)
    if os.path.exists(persist_path):
        return send_file(persist_path, as_attachment=True)
    return "File not found", 404


@app.route('/telegram-webhook', methods=['POST'])
def telegram_webhook():
    """Receive updates from Telegram and handle commands."""
    try:
        update = request.get_json(silent=True) or {}
        msg = update.get('message') or update.get('edited_message') or {}
        chat_id = msg.get('chat', {}).get('id')
        text = msg.get('text', '')
        if chat_id and text:
            tg_handle_command(chat_id, text)
    except Exception as e:
        print(f'[Telegram webhook] error: {e}')
    return jsonify({'ok': True})

def _register_telegram_webhook():
    """Register Railway URL as Telegram webhook on startup."""
    railway_url = os.environ.get('RAILWAY_PUBLIC_DOMAIN') or os.environ.get('RAILWAY_STATIC_URL')
    if not railway_url:
        print('[Telegram] No RAILWAY_PUBLIC_DOMAIN set — webhook not registered (local dev mode)')
        return
    webhook_url = f'https://{railway_url}/telegram-webhook'
    try:
        payload = json.dumps({'url': webhook_url}).encode()
        import urllib.request as _ur
        r = _ur.urlopen(
            _ur.Request(f'{TELEGRAM_API}/setWebhook', data=payload,
                        headers={'Content-Type': 'application/json'}),
            timeout=10
        )
        result = json.loads(r.read())
        print(f'[Telegram] Webhook set to {webhook_url}: {result}')
    except Exception as e:
        print(f'[Telegram] Failed to set webhook: {e}')

if __name__ == '__main__':
    import os
    _register_telegram_webhook()
    port = int(os.environ.get('PORT', 5050))
    print(f"\U0001f3f7\ufe0f  Flipkart Ops Hub running at http://localhost:{port}")
    app.run(host='0.0.0.0', port=port, debug=False)
